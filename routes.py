from flask import render_template, request, redirect, url_for, flash, session, g, jsonify
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from app import app, get_db_connection, allowed_file
import pymysql
import os
import whatsapp_service

# Helper function to get system settings
def get_setting(key, default=None):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT setting_value FROM system_settings WHERE setting_key = %s", (key,))
        result = cursor.fetchone()
        return result['setting_value'] if result else default
    finally:
        cursor.close()
        conn.close()

# Decorator to check if user is logged in
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session and 'employee_id' not in session:
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/')
def index():
    return render_template('index.html')

# User Routes
@app.route('/user/signup', methods=['GET', 'POST'])
def user_signup():
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        
        if not all([username, email, password, confirm_password]):
            flash('All fields are required!', 'danger')
            return redirect(url_for('user_signup'))
        
        if password != confirm_password:
            flash('Passwords do not match!', 'danger')
            return redirect(url_for('user_signup'))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Check if username exists
            cursor.execute("SELECT id FROM users WHERE username = %s", (username,))
            if cursor.fetchone():
                flash('Username already exists!', 'danger')
                return redirect(url_for('user_signup'))
            
            # Check if email exists
            cursor.execute("SELECT id FROM users WHERE email = %s", (email,))
            if cursor.fetchone():
                flash('Email already registered!', 'danger')
                return redirect(url_for('user_signup'))
            
            # Insert new user
            password_hash = generate_password_hash(password)
            cursor.execute(
                "INSERT INTO users (username, email, password_hash) VALUES (%s, %s, %s)",
                (username, email, password_hash)
            )
            conn.commit()
            flash('Account created successfully! Please log in.', 'success')
            return redirect(url_for('user_login'))
        except Exception as e:
            conn.rollback()
            flash('An error occurred. Please try again.', 'danger')
            return redirect(url_for('user_signup'))
        finally:
            cursor.close()
            conn.close()
    
    return render_template('user_signup.html')

@app.route('/user/login', methods=['GET', 'POST'])
def user_login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if not all([username, password]):
            flash('All fields are required!', 'danger')
            return redirect(url_for('user_login'))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute("SELECT * FROM users WHERE username = %s", (username,))
            user = cursor.fetchone()
            
            if user and check_password_hash(user['password_hash'], password):
                session['user_id'] = user['id']
                session['username'] = user['username']
                session['user_type'] = 'user'
                flash(f'Welcome back, {user["username"]}!', 'success')
                return redirect(url_for('user_dashboard'))
            else:
                flash('Invalid username or password!', 'danger')
                return redirect(url_for('user_login'))
        finally:
            cursor.close()
            conn.close()
    
    return render_template('user_login.html')

@app.route('/user/dashboard')
@login_required
def user_dashboard():
    if session.get('user_type') != 'user':
        flash('Unauthorized access!', 'danger')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("SELECT * FROM users WHERE id = %s", (session['user_id'],))
        user = cursor.fetchone()
        return render_template('user_dashboard.html', user=user)
    finally:
        cursor.close()
        conn.close()

# Employee Routes
@app.route('/employee/signup', methods=['GET', 'POST'])
def employee_signup():
    if request.method == 'POST':
        employee_id = request.form.get('employee_id')
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        department = request.form.get('department')
        position = request.form.get('position')
        
        if not all([employee_id, username, email, password, confirm_password]):
            flash('All required fields must be filled!', 'danger')
            return redirect(url_for('employee_signup'))
        
        if password != confirm_password:
            flash('Passwords do not match!', 'danger')
            return redirect(url_for('employee_signup'))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Check if employee_id exists
            cursor.execute("SELECT id FROM employees WHERE employee_id = %s", (employee_id,))
            if cursor.fetchone():
                flash('Employee ID already exists!', 'danger')
                return redirect(url_for('employee_signup'))
            
            # Check if username exists
            cursor.execute("SELECT id FROM employees WHERE username = %s", (username,))
            if cursor.fetchone():
                flash('Username already exists!', 'danger')
                return redirect(url_for('employee_signup'))
            
            # Check if email exists
            cursor.execute("SELECT id FROM employees WHERE email = %s", (email,))
            if cursor.fetchone():
                flash('Email already registered!', 'danger')
                return redirect(url_for('employee_signup'))
            
            # Insert new employee
            password_hash = generate_password_hash(password)
            cursor.execute(
                "INSERT INTO employees (employee_id, username, email, password_hash, department, position) VALUES (%s, %s, %s, %s, %s, %s)",
                (employee_id, username, email, password_hash, department, position)
            )
            conn.commit()
            flash('Employee account created successfully! Please log in.', 'success')
            return redirect(url_for('employee_login'))
        except Exception as e:
            conn.rollback()
            flash('An error occurred. Please try again.', 'danger')
            return redirect(url_for('employee_signup'))
        finally:
            cursor.close()
            conn.close()
    
    return render_template('employee_signup.html')

@app.route('/employee/login', methods=['GET', 'POST'])
def employee_login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if not all([username, password]):
            flash('All fields are required!', 'danger')
            return redirect(url_for('employee_login'))
        
        # Debug: Check tenant context
        print(f"Login attempt - Username: {username}")
        print(f"Has g.tenant_db: {hasattr(g, 'tenant_db')}")
        if hasattr(g, 'tenant_db'):
            print(f"g.tenant_db: {g.tenant_db}")
        print(f"Session tenant_id: {session.get('tenant_id')}")
        print(f"Session tenant_name: {session.get('tenant_name')}")
        
        # Multi-tenant: Try to find user's company first
        multi_tenant_enabled = os.environ.get('MULTI_TENANT_ENABLED', 'true').lower() == 'true'
        if multi_tenant_enabled:
            from tenant_manager import TenantDatabaseManager
            from models_tenant import Company
            
            # Search all tenant databases for this user
            try:
                with TenantDatabaseManager.main_db() as main_conn:
                    with main_conn.cursor() as main_cursor:
                        main_cursor.execute("SELECT * FROM companies WHERE status IN ('active', 'trial')")
                        companies = main_cursor.fetchall()
                        
                        for company in companies:
                            try:
                                # Check if user exists in this tenant's database
                                tenant_conn = TenantDatabaseManager.get_tenant_connection(company['database_name'])
                                tenant_cursor = tenant_conn.cursor()
                                
                                tenant_cursor.execute("SELECT * FROM employees WHERE username = %s", (username,))
                                employee = tenant_cursor.fetchone()
                                
                                if employee and check_password_hash(employee['password_hash'], password):
                                    # Found the user! Set tenant context in session
                                    session['tenant_id'] = company['id']
                                    session['tenant_name'] = company['name']
                                    session['employee_id'] = employee['id']
                                    session['username'] = employee['username']
                                    session['user_type'] = 'employee'
                                    session['role'] = employee.get('role', 'employee')  # Set role for permissions
                                    
                                    print(f"✓ Login successful! Tenant: {company['name']}, User: {employee['username']}")
                                    
                                    flash(f'Welcome back, {employee["username"]}!', 'success')
                                    tenant_cursor.close()
                                    tenant_conn.close()
                                    return redirect(url_for('employee_dashboard'))
                                
                                tenant_cursor.close()
                                tenant_conn.close()
                            except Exception as e:
                                print(f"Error checking tenant {company['name']}: {e}")
                                continue
                
                # User not found in any tenant database
                print(f"✗ Login failed - user not found in any tenant")
                flash('Invalid username or password!', 'danger')
                return redirect(url_for('employee_login'))
                
            except Exception as e:
                print(f"Multi-tenant login error: {e}")
                flash('Login error. Please try again.', 'danger')
                return redirect(url_for('employee_login'))
        else:
            # Single-tenant mode (legacy)
            conn = get_db_connection()
            cursor = conn.cursor()
            
            try:
                cursor.execute("SELECT * FROM employees WHERE username = %s", (username,))
                employee = cursor.fetchone()
                
                if employee and check_password_hash(employee['password_hash'], password):
                    session['employee_id'] = employee['id']
                    session['username'] = employee['username']
                    session['user_type'] = 'employee'
                    session['role'] = employee.get('role', 'employee')  # Set role for permissions
                    flash(f'Welcome back, {employee["username"]}!', 'success')
                    return redirect(url_for('employee_dashboard'))
                else:
                    flash('Invalid username or password!', 'danger')
                    return redirect(url_for('employee_login'))
            finally:
                cursor.close()
                conn.close()
    
    return render_template('employee_login.html')

@app.route('/employee/dashboard')
@login_required
def employee_dashboard():
    if session.get('user_type') != 'employee':
        flash('Unauthorized access!', 'danger')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("SELECT * FROM employees WHERE id = %s", (session['employee_id'],))
        employee = cursor.fetchone()
        
        # Get service notification counts
        cursor.execute("""
            SELECT 
                COUNT(CASE 
                    WHEN (sm.next_service_mileage IS NOT NULL AND fr.latest_mileage IS NOT NULL 
                          AND sm.next_service_mileage <= fr.latest_mileage) 
                    OR (sm.next_service_date IS NOT NULL AND sm.next_service_date <= CURDATE())
                    THEN 1 END) as overdue_count,
                COUNT(CASE 
                    WHEN (sm.next_service_mileage IS NOT NULL AND fr.latest_mileage IS NOT NULL 
                          AND sm.next_service_mileage - fr.latest_mileage BETWEEN 1 AND 1000)
                    OR (sm.next_service_date IS NOT NULL 
                        AND DATEDIFF(sm.next_service_date, CURDATE()) BETWEEN 1 AND 7)
                    THEN 1 END) as due_soon_count
            FROM vehicles v
            LEFT JOIN (
                SELECT vehicle_id, MAX(odometer_reading) as latest_mileage
                FROM fuel_records
                WHERE odometer_reading IS NOT NULL
                GROUP BY vehicle_id
            ) fr ON v.id = fr.vehicle_id
            LEFT JOIN (
                SELECT sm1.vehicle_id, sm1.next_service_mileage, sm1.next_service_date
                FROM service_maintenance sm1
                INNER JOIN (
                    SELECT vehicle_id, MAX(service_date) as max_date
                    FROM service_maintenance
                    GROUP BY vehicle_id
                ) sm2 ON sm1.vehicle_id = sm2.vehicle_id AND sm1.service_date = sm2.max_date
            ) sm ON v.id = sm.vehicle_id
            WHERE (v.status = 'available' OR v.status = 'assigned')
        """)
        notification_counts = cursor.fetchone()
        
        return render_template('employee_dashboard.html', 
                             employee=employee,
                             overdue_count=notification_counts['overdue_count'] or 0,
                             due_soon_count=notification_counts['due_soon_count'] or 0)
    finally:
        cursor.close()
        conn.close()

# Logout Route
@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out successfully.', 'info')
    return redirect(url_for('index'))

# Vehicle Management Routes
@app.route('/vehicles')
@login_required
def vehicles_list():
    if session.get('user_type') != 'employee':
        flash('Only employees can access vehicle management!', 'danger')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Get filter parameters
    type_filter = request.args.get('type', '')
    status_filter = request.args.get('status', '')
    
    try:
        # Build query with filters
        query = """
            SELECT v.*, e.username as added_by_name 
            FROM vehicles v
            LEFT JOIN employees e ON v.added_by = e.id
            WHERE 1=1
        """
        params = []
        
        if type_filter:
            query += " AND v.vehicle_type = %s"
            params.append(type_filter)
        
        if status_filter:
            query += " AND v.status = %s"
            params.append(status_filter)
        
        query += " ORDER BY v.created_at DESC"
        
        cursor.execute(query, params)
        vehicles = cursor.fetchall()
        
        # Get distinct vehicle types for filter
        cursor.execute("SELECT DISTINCT vehicle_type FROM vehicles WHERE vehicle_type IS NOT NULL ORDER BY vehicle_type")
        vehicle_types = cursor.fetchall()
        
        return render_template('vehicles_list.html', 
                             vehicles=vehicles,
                             vehicle_types=vehicle_types,
                             type_filter=type_filter,
                             status_filter=status_filter)
    finally:
        cursor.close()
        conn.close()

@app.route('/vehicles/add', methods=['GET', 'POST'])
@login_required
def add_vehicle():
    if session.get('user_type') != 'employee':
        flash('Only employees can add vehicles!', 'danger')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        vehicle_number = request.form.get('vehicle_number')
        make = request.form.get('make')
        model = request.form.get('model')
        year = request.form.get('year')
        color = request.form.get('color')
        vehicle_type = request.form.get('vehicle_type')
        status = request.form.get('status')
        mileage = request.form.get('mileage')
        last_service_date = request.form.get('last_service_date')
        expected_fuel_consumption = request.form.get('expected_fuel_consumption')
        notes = request.form.get('notes')
        
        if not all([vehicle_number, make, model, year]):
            flash('Vehicle number, make, model, and year are required!', 'danger')
            return redirect(url_for('add_vehicle'))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Check if vehicle number exists
            cursor.execute("SELECT id FROM vehicles WHERE vehicle_number = %s", (vehicle_number,))
            if cursor.fetchone():
                flash('Vehicle number already exists!', 'danger')
                return redirect(url_for('add_vehicle'))
            
            # Insert new vehicle
            cursor.execute("""
                INSERT INTO vehicles (vehicle_number, make, model, year, color, vehicle_type, 
                                    status, mileage, last_service_date, expected_fuel_consumption, notes, added_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (vehicle_number, make, model, year, color, vehicle_type, status, 
                  mileage, last_service_date if last_service_date else None, 
                  expected_fuel_consumption if expected_fuel_consumption else None, notes, session['employee_id']))
            conn.commit()
            flash('Vehicle added successfully!', 'success')
            return redirect(url_for('vehicles_list'))
        except Exception as e:
            conn.rollback()
            flash(f'An error occurred: {str(e)}', 'danger')
            return redirect(url_for('add_vehicle'))
        finally:
            cursor.close()
            conn.close()
    
    return render_template('add_vehicle.html')

@app.route('/vehicles/edit/<int:vehicle_id>', methods=['GET', 'POST'])
@login_required
def edit_vehicle(vehicle_id):
    if session.get('user_type') != 'employee':
        flash('Only employees can edit vehicles!', 'danger')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    if request.method == 'POST':
        vehicle_number = request.form.get('vehicle_number')
        make = request.form.get('make')
        model = request.form.get('model')
        year = request.form.get('year')
        color = request.form.get('color')
        vehicle_type = request.form.get('vehicle_type')
        status = request.form.get('status')
        mileage = request.form.get('mileage')
        last_service_date = request.form.get('last_service_date')
        expected_fuel_consumption = request.form.get('expected_fuel_consumption')
        notes = request.form.get('notes')
        
        if not all([vehicle_number, make, model, year]):
            flash('Vehicle number, make, model, and year are required!', 'danger')
            return redirect(url_for('edit_vehicle', vehicle_id=vehicle_id))
        
        try:
            # Check if vehicle number exists for other vehicles
            cursor.execute("SELECT id FROM vehicles WHERE vehicle_number = %s AND id != %s", 
                         (vehicle_number, vehicle_id))
            if cursor.fetchone():
                flash('Vehicle number already exists!', 'danger')
                return redirect(url_for('edit_vehicle', vehicle_id=vehicle_id))
            
            # Update vehicle
            cursor.execute("""
                UPDATE vehicles 
                SET vehicle_number = %s, make = %s, model = %s, year = %s, color = %s, 
                    vehicle_type = %s, status = %s, mileage = %s, last_service_date = %s, 
                    expected_fuel_consumption = %s, notes = %s
                WHERE id = %s
            """, (vehicle_number, make, model, year, color, vehicle_type, status, 
                  mileage, last_service_date if last_service_date else None, 
                  expected_fuel_consumption if expected_fuel_consumption else None, notes, vehicle_id))
            conn.commit()
            flash('Vehicle updated successfully!', 'success')
            return redirect(url_for('vehicles_list'))
        except Exception as e:
            conn.rollback()
            flash(f'An error occurred: {str(e)}', 'danger')
        finally:
            cursor.close()
            conn.close()
    
    try:
        cursor.execute("SELECT * FROM vehicles WHERE id = %s", (vehicle_id,))
        vehicle = cursor.fetchone()
        if not vehicle:
            flash('Vehicle not found!', 'danger')
            return redirect(url_for('vehicles_list'))
        return render_template('edit_vehicle.html', vehicle=vehicle)
    finally:
        cursor.close()
        conn.close()

@app.route('/vehicles/delete/<int:vehicle_id>', methods=['POST'])
@login_required
def delete_vehicle(vehicle_id):
    if session.get('user_type') != 'employee':
        flash('Only employees can delete vehicles!', 'danger')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("DELETE FROM vehicles WHERE id = %s", (vehicle_id,))
        conn.commit()
        flash('Vehicle deleted successfully!', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'An error occurred: {str(e)}', 'danger')
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('vehicles_list'))

@app.route('/vehicles/view/<int:vehicle_id>')
@login_required
def view_vehicle(vehicle_id):
    if session.get('user_type') != 'employee':
        flash('Only employees can view vehicle details!', 'danger')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT v.*, e.username as added_by_name, e.employee_id as added_by_emp_id
            FROM vehicles v
            LEFT JOIN employees e ON v.added_by = e.id
            WHERE v.id = %s
        """, (vehicle_id,))
        vehicle = cursor.fetchone()
        
        if not vehicle:
            flash('Vehicle not found!', 'danger')
            return redirect(url_for('vehicles_list'))
        
        return render_template('view_vehicle.html', vehicle=vehicle)
    finally:
        cursor.close()
        conn.close()

# Vehicle Assignment Routes
@app.route('/assignments')
@login_required
def assignments_list():
    if session.get('user_type') != 'employee':
        flash('Only employees can access assignments!', 'danger')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Get filter parameters
    vehicle_filter = request.args.get('vehicle', '')
    status_filter = request.args.get('status', '')
    
    try:
        # Build query with filters
        query = """
            SELECT va.*, 
                   v.vehicle_number, v.make, v.model,
                   e.username as driver_name, e.employee_id as driver_emp_id,
                   e2.username as assigned_by_name
            FROM vehicle_assignments va
            JOIN vehicles v ON va.vehicle_id = v.id
            JOIN employees e ON va.employee_id = e.id
            LEFT JOIN employees e2 ON va.assigned_by = e2.id
            WHERE 1=1
        """
        params = []
        
        if vehicle_filter:
            query += " AND va.vehicle_id = %s"
            params.append(vehicle_filter)
        
        if status_filter:
            query += " AND va.status = %s"
            params.append(status_filter)
        
        query += " ORDER BY va.assignment_date DESC"
        
        cursor.execute(query, params)
        assignments = cursor.fetchall()
        
        # Get all vehicles for filter dropdown
        cursor.execute("""
            SELECT DISTINCT v.id, v.vehicle_number, v.make, v.model
            FROM vehicles v
            INNER JOIN vehicle_assignments va ON v.id = va.vehicle_id
            ORDER BY v.vehicle_number
        """)
        vehicles = cursor.fetchall()
        
        return render_template('assignments_list.html', 
                             assignments=assignments, 
                             vehicles=vehicles,
                             vehicle_filter=vehicle_filter,
                             status_filter=status_filter)
    finally:
        cursor.close()
        conn.close()

@app.route('/assignments/assign', methods=['GET', 'POST'])
@login_required
def assign_vehicle():
    if session.get('user_type') != 'employee':
        flash('Only employees can assign vehicles!', 'danger')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    if request.method == 'POST':
        vehicle_id = request.form.get('vehicle_id')
        employee_id = request.form.get('employee_id')
        purpose = request.form.get('purpose')
        notes = request.form.get('notes')
        mileage_at_assignment = request.form.get('mileage_at_assignment')
        
        if not all([vehicle_id, employee_id]):
            flash('Vehicle and Employee are required!', 'danger')
            return redirect(url_for('assign_vehicle'))
        
        try:
            # Check if vehicle is available
            cursor.execute("SELECT status, mileage FROM vehicles WHERE id = %s", (vehicle_id,))
            vehicle = cursor.fetchone()
            
            if not vehicle:
                flash('Vehicle not found!', 'danger')
                return redirect(url_for('assign_vehicle'))
            
            if vehicle['status'] != 'available':
                flash('Vehicle is not available for assignment!', 'warning')
                return redirect(url_for('assign_vehicle'))
            
            # Use current vehicle mileage if not provided
            if not mileage_at_assignment:
                mileage_at_assignment = vehicle['mileage']
            
            # Check if employee has active assignment
            cursor.execute("""
                SELECT COUNT(*) as count FROM vehicle_assignments 
                WHERE employee_id = %s AND status = 'active'
            """, (employee_id,))
            result = cursor.fetchone()
            
            if result['count'] > 0:
                flash('This employee already has an active vehicle assignment!', 'warning')
                return redirect(url_for('assign_vehicle'))
            
            # Create assignment
            cursor.execute("""
                INSERT INTO vehicle_assignments 
                (vehicle_id, employee_id, assigned_by, purpose, notes, status, mileage_at_assignment)
                VALUES (%s, %s, %s, %s, %s, 'active', %s)
            """, (vehicle_id, employee_id, session['employee_id'], purpose, notes, mileage_at_assignment))
            
            # Update vehicle status and mileage
            cursor.execute("""
                UPDATE vehicles SET status = 'in_use', mileage = %s WHERE id = %s
            """, (mileage_at_assignment, vehicle_id))
            
            conn.commit()
            flash('Vehicle assigned successfully!', 'success')
            
            # Send WhatsApp notification to assigned employee
            try:
                cursor.execute("""
                    SELECT ns.vehicle_assignment_whatsapp, ns.whatsapp_number, e.username
                    FROM notification_settings ns
                    JOIN employees e ON ns.employee_id = e.id
                    WHERE ns.employee_id = %s AND ns.vehicle_assignment_whatsapp = TRUE
                """, (employee_id,))
                notif_pref = cursor.fetchone()
                
                if notif_pref and notif_pref['whatsapp_number']:
                    cursor.execute("""
                        SELECT v.vehicle_number, v.make, v.model
                        FROM vehicles v WHERE v.id = %s
                    """, (vehicle_id,))
                    vehicle_info = cursor.fetchone()
                    
                    vehicle_details = {
                        'make': vehicle_info['make'],
                        'model': vehicle_info['model'],
                        'assignment_date': 'Today',
                        'purpose': purpose if purpose else 'General Use'
                    }
                    
                    whatsapp_service.send_vehicle_assignment_alert(
                        vehicle_info['vehicle_number'],
                        vehicle_details,
                        notif_pref['username'],
                        notif_pref['whatsapp_number']
                    )
            except Exception as e:
                print(f"WhatsApp notification error: {e}")
            
            return redirect(url_for('assignments_list'))
        except Exception as e:
            conn.rollback()
            flash(f'An error occurred: {str(e)}', 'danger')
            return redirect(url_for('assign_vehicle'))
        finally:
            cursor.close()
            conn.close()
    
    try:
        # Get available vehicles
        cursor.execute("""
            SELECT id, vehicle_number, make, model, year, mileage 
            FROM vehicles 
            WHERE status = 'available'
            ORDER BY vehicle_number
        """)
        vehicles = cursor.fetchall()
        
        # Get all employees
        cursor.execute("""
            SELECT id, employee_id, username, position 
            FROM employees 
            ORDER BY username
        """)
        employees = cursor.fetchall()
        
        return render_template('assign_vehicle.html', vehicles=vehicles, employees=employees)
    finally:
        cursor.close()
        conn.close()

@app.route('/assignments/return/<int:assignment_id>', methods=['GET', 'POST'])
@login_required
def return_vehicle(assignment_id):
    if session.get('user_type') != 'employee':
        flash('Only employees can process vehicle returns!', 'danger')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    if request.method == 'POST':
        mileage_at_return = request.form.get('mileage_at_return')
        
        if not mileage_at_return:
            flash('Mileage at return is required!', 'danger')
            return redirect(url_for('return_vehicle', assignment_id=assignment_id))
        
        try:
            # Get assignment details
            cursor.execute("""
                SELECT vehicle_id, status, mileage_at_assignment FROM vehicle_assignments WHERE id = %s
            """, (assignment_id,))
            assignment = cursor.fetchone()
            
            if not assignment:
                flash('Assignment not found!', 'danger')
                return redirect(url_for('assignments_list'))
            
            if assignment['status'] != 'active':
                flash('This assignment is not active!', 'warning')
                return redirect(url_for('assignments_list'))
            
            # Validate mileage
            if assignment['mileage_at_assignment'] and int(mileage_at_return) < assignment['mileage_at_assignment']:
                flash('Return mileage cannot be less than assignment mileage!', 'danger')
                return redirect(url_for('return_vehicle', assignment_id=assignment_id))
            
            # Update assignment
            cursor.execute("""
                UPDATE vehicle_assignments 
                SET status = 'completed', return_date = NOW(), mileage_at_return = %s
                WHERE id = %s
            """, (mileage_at_return, assignment_id))
            
            # Update vehicle status and mileage
            cursor.execute("""
                UPDATE vehicles SET status = 'available', mileage = %s WHERE id = %s
            """, (mileage_at_return, assignment['vehicle_id']))
            
            conn.commit()
            flash('Vehicle returned successfully!', 'success')
            return redirect(url_for('assignments_list'))
        except Exception as e:
            conn.rollback()
            flash(f'An error occurred: {str(e)}', 'danger')
            return redirect(url_for('return_vehicle', assignment_id=assignment_id))
        finally:
            cursor.close()
            conn.close()
    
    # GET request - show return form
    try:
        cursor.execute("""
            SELECT va.*, 
                   v.vehicle_number, v.make, v.model, v.mileage as current_mileage,
                   e.username as driver_name
            FROM vehicle_assignments va
            JOIN vehicles v ON va.vehicle_id = v.id
            JOIN employees e ON va.employee_id = e.id
            WHERE va.id = %s
        """, (assignment_id,))
        assignment = cursor.fetchone()
        
        if not assignment:
            flash('Assignment not found!', 'danger')
            return redirect(url_for('assignments_list'))
        
        if assignment['status'] != 'active':
            flash('This assignment is not active!', 'warning')
            return redirect(url_for('assignments_list'))
        
        return render_template('return_vehicle.html', assignment=assignment)
    finally:
        cursor.close()
        conn.close()

@app.route('/assignments/view/<int:assignment_id>')
@login_required
def view_assignment(assignment_id):
    if session.get('user_type') != 'employee':
        flash('Only employees can view assignments!', 'danger')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT va.*, 
                   v.vehicle_number, v.make, v.model, v.year, v.color, v.vehicle_type,
                   e.username as driver_name, e.employee_id as driver_emp_id, 
                   e.email as driver_email, e.position as driver_position,
                   e2.username as assigned_by_name, e2.employee_id as assigned_by_emp_id
            FROM vehicle_assignments va
            JOIN vehicles v ON va.vehicle_id = v.id
            JOIN employees e ON va.employee_id = e.id
            LEFT JOIN employees e2 ON va.assigned_by = e2.id
            WHERE va.id = %s
        """, (assignment_id,))
        assignment = cursor.fetchone()
        
        if not assignment:
            flash('Assignment not found!', 'danger')
            return redirect(url_for('assignments_list'))
        
        return render_template('view_assignment.html', assignment=assignment)
    finally:
        cursor.close()
        conn.close()

@app.route('/my-assignments')
@login_required
def my_assignments():
    if session.get('user_type') != 'employee':
        flash('Only employees can view their assignments!', 'danger')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT va.*, 
                   v.vehicle_number, v.make, v.model, v.year, v.color,
                   e2.username as assigned_by_name
            FROM vehicle_assignments va
            JOIN vehicles v ON va.vehicle_id = v.id
            LEFT JOIN employees e2 ON va.assigned_by = e2.id
            WHERE va.employee_id = %s
            ORDER BY va.assignment_date DESC
        """, (session['employee_id'],))
        assignments = cursor.fetchall()
        return render_template('my_assignments.html', assignments=assignments)
    finally:
        cursor.close()
        conn.close()

# Fuel Tracking Routes
@app.route('/fuel-records')
def fuel_records():
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Get filter parameters
        vehicle_filter = request.args.get('vehicle_filter', '')
        employee_filter = request.args.get('employee_filter', '')
        
        # Build query
        query = """
            SELECT fr.*, 
                   v.vehicle_number, v.make, v.model,
                   e.username as employee_name
            FROM fuel_records fr
            JOIN vehicles v ON fr.vehicle_id = v.id
            JOIN employees e ON fr.employee_id = e.id
            WHERE 1=1
        """
        params = []
        
        if vehicle_filter:
            query += " AND fr.vehicle_id = %s"
            params.append(vehicle_filter)
        
        if employee_filter:
            query += " AND fr.employee_id = %s"
            params.append(employee_filter)
        
        query += " ORDER BY fr.fuel_date DESC"
        
        cursor.execute(query, params)
        fuel_records = cursor.fetchall()
        
        # Get vehicles and employees for filters
        cursor.execute("SELECT id, vehicle_number, make, model FROM vehicles ORDER BY vehicle_number")
        vehicles = cursor.fetchall()
        
        cursor.execute("SELECT id, username FROM employees ORDER BY username")
        employees = cursor.fetchall()
        
        return render_template('fuel_records.html', 
                             fuel_records=fuel_records, 
                             vehicles=vehicles,
                             employees=employees,
                             vehicle_filter=vehicle_filter,
                             employee_filter=employee_filter)
    finally:
        cursor.close()
        conn.close()

@app.route('/add-fuel-record', methods=['GET', 'POST'])
def add_fuel_record():
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    if request.method == 'POST':
        vehicle_id = request.form.get('vehicle_id')
        fuel_amount = request.form.get('fuel_amount')
        fuel_cost = request.form.get('fuel_cost')
        odometer_reading = request.form.get('odometer_reading')
        fuel_type = request.form.get('fuel_type')
        station_name = request.form.get('station_name')
        notes = request.form.get('notes')
        
        # Handle file upload
        receipt_path = None
        if 'receipt' in request.files:
            file = request.files['receipt']
            if file and file.filename != '':
                if allowed_file(file.filename):
                    from werkzeug.utils import secure_filename
                    import uuid
                    # Create unique filename
                    filename = secure_filename(file.filename)
                    unique_filename = f"{uuid.uuid4()}_{filename}"
                    file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
                    file.save(file_path)
                    receipt_path = f"uploads/fuel_receipts/{unique_filename}"
                else:
                    flash('Invalid file type! Allowed types: PNG, JPG, JPEG, PDF', 'danger')
                    return redirect(url_for('add_fuel_record'))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("""
                INSERT INTO fuel_records 
                (vehicle_id, employee_id, fuel_amount, fuel_cost, odometer_reading, 
                 fuel_type, station_name, receipt_path, notes)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (vehicle_id, session['employee_id'], fuel_amount, fuel_cost, 
                  odometer_reading, fuel_type, station_name, receipt_path, notes))
            conn.commit()
            flash('Fuel record added successfully!', 'success')
            
            # Check fuel expense threshold and send WhatsApp notification
            try:
                threshold = float(get_setting('fuel_price_alert_threshold', 5.0))
                price_per_liter = float(fuel_cost) / float(fuel_amount)
                
                if price_per_liter > threshold:
                    cursor.execute("""
                        SELECT ns.fuel_expense_alert_whatsapp, ns.whatsapp_number, v.vehicle_number
                        FROM notification_settings ns
                        JOIN vehicles v ON v.id = %s
                        WHERE ns.employee_id = %s AND ns.fuel_expense_alert_whatsapp = TRUE
                    """, (vehicle_id, session['employee_id']))
                    notif_pref = cursor.fetchone()
                    
                    if notif_pref and notif_pref['whatsapp_number']:
                        whatsapp_service.send_fuel_expense_alert(
                            notif_pref['vehicle_number'],
                            fuel_cost,
                            threshold,
                            notif_pref['whatsapp_number']
                        )
            except Exception as e:
                print(f"WhatsApp notification error: {e}")
            
            return redirect(url_for('fuel_records'))
        finally:
            cursor.close()
            conn.close()
    
    # GET request - show form
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT id, vehicle_number, make, model, mileage 
            FROM vehicles 
            WHERE status = 'available' OR status = 'in_use'
            ORDER BY vehicle_number
        """)
        vehicles = cursor.fetchall()
        return render_template('add_fuel_record.html', vehicles=vehicles)
    finally:
        cursor.close()
        conn.close()

@app.route('/view-fuel-record/<int:record_id>')
def view_fuel_record(record_id):
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT fr.*, 
                   v.vehicle_number, v.make, v.model,
                   e.username as employee_name
            FROM fuel_records fr
            JOIN vehicles v ON fr.vehicle_id = v.id
            JOIN employees e ON fr.employee_id = e.id
            WHERE fr.id = %s
        """, (record_id,))
        record = cursor.fetchone()
        
        if not record:
            flash('Fuel record not found!', 'danger')
            return redirect(url_for('fuel_records'))
        
        return render_template('view_fuel_record.html', record=record)
    finally:
        cursor.close()
        conn.close()

@app.route('/delete-fuel-record/<int:record_id>')
def delete_fuel_record(record_id):
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Get receipt path before deleting
        cursor.execute("SELECT receipt_path FROM fuel_records WHERE id = %s", (record_id,))
        record = cursor.fetchone()
        
        if record and record['receipt_path']:
            # Delete file from filesystem
            file_path = os.path.join('static', record['receipt_path'])
            if os.path.exists(file_path):
                os.remove(file_path)
        
        cursor.execute("DELETE FROM fuel_records WHERE id = %s", (record_id,))
        conn.commit()
        flash('Fuel record deleted successfully!', 'success')
        return redirect(url_for('fuel_records'))
    finally:
        cursor.close()
        conn.close()

@app.route('/fuel-reports')
def fuel_reports():
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    from datetime import datetime, timedelta
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Get filter parameters
        report_type = request.args.get('report_type', 'monthly')  # weekly, monthly, quarterly, custom
        vehicle_filter = request.args.get('vehicle_filter', '')
        driver_filter = request.args.get('driver_filter', '')
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        
        # Calculate date ranges based on report type
        today = datetime.now().date()
        
        if report_type == 'weekly':
            start_date = (today - timedelta(days=today.weekday())).strftime('%Y-%m-%d')
            end_date = (today + timedelta(days=(6 - today.weekday()))).strftime('%Y-%m-%d')
        elif report_type == 'monthly':
            start_date = today.replace(day=1).strftime('%Y-%m-%d')
            # Get last day of month
            if today.month == 12:
                next_month = today.replace(year=today.year + 1, month=1, day=1)
            else:
                next_month = today.replace(month=today.month + 1, day=1)
            end_date = (next_month - timedelta(days=1)).strftime('%Y-%m-%d')
        elif report_type == 'quarterly':
            quarter = (today.month - 1) // 3
            start_month = quarter * 3 + 1
            start_date = today.replace(month=start_month, day=1).strftime('%Y-%m-%d')
            # Get last day of quarter
            end_month = start_month + 2
            if end_month == 12:
                end_date = today.replace(month=12, day=31).strftime('%Y-%m-%d')
            else:
                next_month = today.replace(month=end_month + 1, day=1)
                end_date = (next_month - timedelta(days=1)).strftime('%Y-%m-%d')
        # For custom, use provided dates or default to current month
        elif not start_date or not end_date:
            start_date = today.replace(day=1).strftime('%Y-%m-%d')
            if today.month == 12:
                next_month = today.replace(year=today.year + 1, month=1, day=1)
            else:
                next_month = today.replace(month=today.month + 1, day=1)
            end_date = (next_month - timedelta(days=1)).strftime('%Y-%m-%d')
        
        # Build query
        query = """
            SELECT fr.*, 
                   v.vehicle_number, v.make, v.model, v.id as vehicle_id,
                   e.username as driver_name, e.id as driver_id
            FROM fuel_records fr
            JOIN vehicles v ON fr.vehicle_id = v.id
            JOIN employees e ON fr.employee_id = e.id
            WHERE fr.fuel_date BETWEEN %s AND %s
        """
        params = [start_date, end_date]
        
        if vehicle_filter:
            query += " AND fr.vehicle_id = %s"
            params.append(vehicle_filter)
        
        if driver_filter:
            query += " AND fr.employee_id = %s"
            params.append(driver_filter)
        
        query += " ORDER BY v.vehicle_number, fr.fuel_date ASC"
        
        cursor.execute(query, params)
        fuel_records = cursor.fetchall()
        
        # Calculate distance and consumption for each record
        # Group by vehicle to track previous odometer
        vehicle_previous = {}
        for record in fuel_records:
            vid = record['vehicle_id']
            record['distance_since_last'] = None
            record['consumption'] = None
            
            if record['odometer_reading']:
                if vid in vehicle_previous and vehicle_previous[vid]['odometer']:
                    distance = record['odometer_reading'] - vehicle_previous[vid]['odometer']
                    if distance > 0:
                        record['distance_since_last'] = distance
                        # Calculate consumption in L/100km
                        record['consumption'] = (float(record['fuel_amount']) / distance) * 100
                
                vehicle_previous[vid] = {
                    'odometer': record['odometer_reading'],
                    'litres': record['fuel_amount']
                }
        
        # Reverse for display (newest first) - convert tuple to list first
        fuel_records = list(fuel_records)
        fuel_records.reverse()
        
        # Calculate summary statistics
        summary = {
            'total_records': len(fuel_records),
            'total_litres': sum(float(record['fuel_amount']) for record in fuel_records),
            'total_cost': sum(float(record['fuel_cost']) for record in fuel_records),
            'total_odometer': 0,
            'avg_cost_per_litre': 0,
            'avg_litres_per_fill': 0
        }
        
        if summary['total_records'] > 0:
            summary['avg_litres_per_fill'] = summary['total_litres'] / summary['total_records']
            if summary['total_litres'] > 0:
                summary['avg_cost_per_litre'] = summary['total_cost'] / summary['total_litres']
        
        # Group by vehicle
        vehicles_summary = {}
        for record in fuel_records:
            vid = record['vehicle_id']
            if vid not in vehicles_summary:
                vehicles_summary[vid] = {
                    'vehicle_number': record['vehicle_number'],
                    'make': record['make'],
                    'model': record['model'],
                    'total_litres': 0,
                    'total_cost': 0,
                    'fill_count': 0,
                    'total_km': 0,
                    'avg_consumption': 0,
                    'records': []
                }
            vehicles_summary[vid]['total_litres'] += float(record['fuel_amount'])
            vehicles_summary[vid]['total_cost'] += float(record['fuel_cost'])
            vehicles_summary[vid]['fill_count'] += 1
            vehicles_summary[vid]['records'].append(record)
        
        # Calculate km traveled and fuel consumption per vehicle
        for vid, vdata in vehicles_summary.items():
            # Sort records by date to calculate distance
            sorted_records = sorted(vdata['records'], key=lambda x: x['fuel_date'])
            if len(sorted_records) >= 2:
                # Calculate total km from first to last odometer reading
                first_odometer = sorted_records[0]['odometer_reading'] if sorted_records[0]['odometer_reading'] else 0
                last_odometer = sorted_records[-1]['odometer_reading'] if sorted_records[-1]['odometer_reading'] else 0
                vdata['total_km'] = last_odometer - first_odometer
                
                # Calculate average consumption (litres per 100km)
                if vdata['total_km'] > 0:
                    vdata['avg_consumption'] = (vdata['total_litres'] / vdata['total_km']) * 100
        
        # Group by driver
        drivers_summary = {}
        for record in fuel_records:
            did = record['driver_id']
            if did not in drivers_summary:
                drivers_summary[did] = {
                    'driver_name': record['driver_name'],
                    'total_litres': 0,
                    'total_cost': 0,
                    'fill_count': 0,
                    'total_km': 0,
                    'avg_consumption': 0,
                    'records': []
                }
            drivers_summary[did]['total_litres'] += float(record['fuel_amount'])
            drivers_summary[did]['total_cost'] += float(record['fuel_cost'])
            drivers_summary[did]['fill_count'] += 1
            drivers_summary[did]['records'].append(record)
        
        # Calculate km traveled and fuel consumption per driver
        for did, ddata in drivers_summary.items():
            # Sort records by date
            sorted_records = sorted(ddata['records'], key=lambda x: x['fuel_date'])
            if len(sorted_records) >= 2:
                first_odometer = sorted_records[0]['odometer_reading'] if sorted_records[0]['odometer_reading'] else 0
                last_odometer = sorted_records[-1]['odometer_reading'] if sorted_records[-1]['odometer_reading'] else 0
                ddata['total_km'] = last_odometer - first_odometer
                
                if ddata['total_km'] > 0:
                    ddata['avg_consumption'] = (ddata['total_litres'] / ddata['total_km']) * 100
        
        # Get all vehicles and drivers for filters
        cursor.execute("SELECT id, vehicle_number, make, model FROM vehicles ORDER BY vehicle_number")
        vehicles = cursor.fetchall()
        
        cursor.execute("SELECT id, username FROM employees ORDER BY username")
        employees = cursor.fetchall()
        
        return render_template('fuel_reports.html',
                             fuel_records=fuel_records,
                             summary=summary,
                             vehicles_summary=vehicles_summary,
                             drivers_summary=drivers_summary,
                             vehicles=vehicles,
                             employees=employees,
                             report_type=report_type,
                             vehicle_filter=vehicle_filter,
                             driver_filter=driver_filter,
                             start_date=start_date,
                             end_date=end_date)
    finally:
        cursor.close()
        conn.close()

@app.route('/fuel-reports/print')
def fuel_reports_print():
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    from datetime import datetime, timedelta
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Get filter parameters
        report_type = request.args.get('report_type', 'monthly')
        vehicle_filter = request.args.get('vehicle_filter', '')
        driver_filter = request.args.get('driver_filter', '')
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        
        # Calculate date ranges based on report type
        today = datetime.now().date()
        
        if report_type == 'weekly':
            start_date = (today - timedelta(days=today.weekday())).strftime('%Y-%m-%d')
            end_date = (today + timedelta(days=(6 - today.weekday()))).strftime('%Y-%m-%d')
        elif report_type == 'monthly':
            start_date = today.replace(day=1).strftime('%Y-%m-%d')
            if today.month == 12:
                next_month = today.replace(year=today.year + 1, month=1, day=1)
            else:
                next_month = today.replace(month=today.month + 1, day=1)
            end_date = (next_month - timedelta(days=1)).strftime('%Y-%m-%d')
        elif report_type == 'quarterly':
            quarter = (today.month - 1) // 3
            start_month = quarter * 3 + 1
            start_date = today.replace(month=start_month, day=1).strftime('%Y-%m-%d')
            end_month = start_month + 2
            if end_month == 12:
                end_date = today.replace(month=12, day=31).strftime('%Y-%m-%d')
            else:
                next_month = today.replace(month=end_month + 1, day=1)
                end_date = (next_month - timedelta(days=1)).strftime('%Y-%m-%d')
        elif not start_date or not end_date:
            start_date = today.replace(day=1).strftime('%Y-%m-%d')
            if today.month == 12:
                next_month = today.replace(year=today.year + 1, month=1, day=1)
            else:
                next_month = today.replace(month=today.month + 1, day=1)
            end_date = (next_month - timedelta(days=1)).strftime('%Y-%m-%d')
        
        # Build query
        query = """
            SELECT fr.*, 
                   v.vehicle_number, v.make, v.model, v.id as vehicle_id,
                   e.username as driver_name, e.id as driver_id
            FROM fuel_records fr
            JOIN vehicles v ON fr.vehicle_id = v.id
            JOIN employees e ON fr.employee_id = e.id
            WHERE fr.fuel_date BETWEEN %s AND %s
        """
        params = [start_date, end_date]
        
        if vehicle_filter:
            query += " AND fr.vehicle_id = %s"
            params.append(vehicle_filter)
        
        if driver_filter:
            query += " AND fr.employee_id = %s"
            params.append(driver_filter)
        
        query += " ORDER BY fr.fuel_date DESC, v.vehicle_number"
        
        cursor.execute(query, params)
        fuel_records = cursor.fetchall()
        
        # Calculate summary statistics
        summary = {
            'total_records': len(fuel_records),
            'total_litres': sum(float(record['fuel_amount']) for record in fuel_records),
            'total_cost': sum(float(record['fuel_cost']) for record in fuel_records),
            'avg_cost_per_litre': 0,
            'avg_litres_per_fill': 0
        }
        
        if summary['total_records'] > 0:
            summary['avg_litres_per_fill'] = summary['total_litres'] / summary['total_records']
            if summary['total_litres'] > 0:
                summary['avg_cost_per_litre'] = summary['total_cost'] / summary['total_litres']
        
        # Group by vehicle
        vehicles_summary = {}
        for record in fuel_records:
            vid = record['vehicle_id']
            if vid not in vehicles_summary:
                vehicles_summary[vid] = {
                    'vehicle_number': record['vehicle_number'],
                    'make': record['make'],
                    'model': record['model'],
                    'total_litres': 0,
                    'total_cost': 0,
                    'fill_count': 0,
                    'total_km': 0,
                    'avg_consumption': 0,
                    'records': []
                }
            vehicles_summary[vid]['total_litres'] += float(record['fuel_amount'])
            vehicles_summary[vid]['total_cost'] += float(record['fuel_cost'])
            vehicles_summary[vid]['fill_count'] += 1
            vehicles_summary[vid]['records'].append(record)
        
        # Calculate km traveled and fuel consumption per vehicle
        for vid, vdata in vehicles_summary.items():
            sorted_records = sorted(vdata['records'], key=lambda x: x['fuel_date'])
            if len(sorted_records) >= 2:
                first_odometer = sorted_records[0]['odometer_reading'] if sorted_records[0]['odometer_reading'] else 0
                last_odometer = sorted_records[-1]['odometer_reading'] if sorted_records[-1]['odometer_reading'] else 0
                vdata['total_km'] = last_odometer - first_odometer
                
                if vdata['total_km'] > 0:
                    vdata['avg_consumption'] = (vdata['total_litres'] / vdata['total_km']) * 100
        
        # Group by driver
        drivers_summary = {}
        for record in fuel_records:
            did = record['driver_id']
            if did not in drivers_summary:
                drivers_summary[did] = {
                    'driver_name': record['driver_name'],
                    'total_litres': 0,
                    'total_cost': 0,
                    'fill_count': 0,
                    'total_km': 0,
                    'avg_consumption': 0,
                    'records': []
                }
            drivers_summary[did]['total_litres'] += float(record['fuel_amount'])
            drivers_summary[did]['total_cost'] += float(record['fuel_cost'])
            drivers_summary[did]['fill_count'] += 1
            drivers_summary[did]['records'].append(record)
        
        # Calculate km traveled and fuel consumption per driver
        for did, ddata in drivers_summary.items():
            sorted_records = sorted(ddata['records'], key=lambda x: x['fuel_date'])
            if len(sorted_records) >= 2:
                first_odometer = sorted_records[0]['odometer_reading'] if sorted_records[0]['odometer_reading'] else 0
                last_odometer = sorted_records[-1]['odometer_reading'] if sorted_records[-1]['odometer_reading'] else 0
                ddata['total_km'] = last_odometer - first_odometer
                
                if ddata['total_km'] > 0:
                    ddata['avg_consumption'] = (ddata['total_litres'] / ddata['total_km']) * 100
        
        # Get print settings
        cursor.execute("SELECT * FROM print_settings LIMIT 1")
        print_config = cursor.fetchone()
        
        return render_template('fuel_reports_print.html',
                             fuel_records=fuel_records,
                             summary=summary,
                             vehicles_summary=vehicles_summary,
                             drivers_summary=drivers_summary,
                             report_type=report_type,
                             start_date=start_date,
                             end_date=end_date,
                             print_config=print_config)
    finally:
        cursor.close()
        conn.close()

@app.route('/fuel-reports/export')
def fuel_reports_export():
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    from datetime import datetime, timedelta
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import send_file
    import io
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Get filter parameters
        report_type = request.args.get('report_type', 'monthly')
        vehicle_filter = request.args.get('vehicle_filter', '')
        driver_filter = request.args.get('driver_filter', '')
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        
        # Calculate date ranges based on report type
        today = datetime.now().date()
        
        if report_type == 'weekly':
            start_date = (today - timedelta(days=today.weekday())).strftime('%Y-%m-%d')
            end_date = (today + timedelta(days=(6 - today.weekday()))).strftime('%Y-%m-%d')
        elif report_type == 'monthly':
            start_date = today.replace(day=1).strftime('%Y-%m-%d')
            if today.month == 12:
                next_month = today.replace(year=today.year + 1, month=1, day=1)
            else:
                next_month = today.replace(month=today.month + 1, day=1)
            end_date = (next_month - timedelta(days=1)).strftime('%Y-%m-%d')
        elif report_type == 'quarterly':
            quarter = (today.month - 1) // 3
            start_month = quarter * 3 + 1
            start_date = today.replace(month=start_month, day=1).strftime('%Y-%m-%d')
            end_month = start_month + 2
            if end_month == 12:
                end_date = today.replace(month=12, day=31).strftime('%Y-%m-%d')
            else:
                next_month = today.replace(month=end_month + 1, day=1)
                end_date = (next_month - timedelta(days=1)).strftime('%Y-%m-%d')
        elif not start_date or not end_date:
            start_date = today.replace(day=1).strftime('%Y-%m-%d')
            if today.month == 12:
                next_month = today.replace(year=today.year + 1, month=1, day=1)
            else:
                next_month = today.replace(month=today.month + 1, day=1)
            end_date = (next_month - timedelta(days=1)).strftime('%Y-%m-%d')
        
        # Build query
        query = """
            SELECT fr.*, 
                   v.vehicle_number, v.make, v.model, v.id as vehicle_id,
                   e.username as driver_name, e.id as driver_id
            FROM fuel_records fr
            JOIN vehicles v ON fr.vehicle_id = v.id
            JOIN employees e ON fr.employee_id = e.id
            WHERE fr.fuel_date BETWEEN %s AND %s
        """
        params = [start_date, end_date]
        
        if vehicle_filter:
            query += " AND fr.vehicle_id = %s"
            params.append(vehicle_filter)
        
        if driver_filter:
            query += " AND fr.employee_id = %s"
            params.append(driver_filter)
        
        query += " ORDER BY v.vehicle_number, fr.fuel_date ASC"
        
        cursor.execute(query, params)
        fuel_records = cursor.fetchall()
        
        # Calculate distance and consumption for each record
        vehicle_previous = {}
        for record in fuel_records:
            vid = record['vehicle_id']
            record['distance_since_last'] = None
            record['consumption'] = None
            
            if record['odometer_reading']:
                if vid in vehicle_previous and vehicle_previous[vid]['odometer']:
                    distance = record['odometer_reading'] - vehicle_previous[vid]['odometer']
                    if distance > 0:
                        record['distance_since_last'] = distance
                        record['consumption'] = (float(record['fuel_amount']) / distance) * 100
                
                vehicle_previous[vid] = {
                    'odometer': record['odometer_reading'],
                    'litres': record['fuel_amount']
                }
        
        # Reverse for display
        fuel_records.reverse()
        
        # Calculate summary statistics
        summary = {
            'total_records': len(fuel_records),
            'total_litres': sum(float(record['fuel_amount']) for record in fuel_records),
            'total_cost': sum(float(record['fuel_cost']) for record in fuel_records),
            'avg_cost_per_litre': 0,
            'avg_litres_per_fill': 0
        }
        
        if summary['total_records'] > 0:
            summary['avg_litres_per_fill'] = summary['total_litres'] / summary['total_records']
            if summary['total_litres'] > 0:
                summary['avg_cost_per_litre'] = summary['total_cost'] / summary['total_litres']
        
        # Group by vehicle
        vehicles_summary = {}
        for record in fuel_records:
            vid = record['vehicle_id']
            if vid not in vehicles_summary:
                vehicles_summary[vid] = {
                    'vehicle_number': record['vehicle_number'],
                    'make': record['make'],
                    'model': record['model'],
                    'total_litres': 0,
                    'total_cost': 0,
                    'fill_count': 0,
                    'total_km': 0,
                    'avg_consumption': 0,
                    'records': []
                }
            vehicles_summary[vid]['total_litres'] += float(record['fuel_amount'])
            vehicles_summary[vid]['total_cost'] += float(record['fuel_cost'])
            vehicles_summary[vid]['fill_count'] += 1
            vehicles_summary[vid]['records'].append(record)
        
        # Calculate km traveled per vehicle
        for vid, vdata in vehicles_summary.items():
            sorted_records = sorted(vdata['records'], key=lambda x: x['fuel_date'])
            if len(sorted_records) >= 2:
                first_odometer = sorted_records[0]['odometer_reading'] if sorted_records[0]['odometer_reading'] else 0
                last_odometer = sorted_records[-1]['odometer_reading'] if sorted_records[-1]['odometer_reading'] else 0
                vdata['total_km'] = last_odometer - first_odometer
                
                if vdata['total_km'] > 0:
                    vdata['avg_consumption'] = (vdata['total_litres'] / vdata['total_km']) * 100
        
        # Create Excel workbook
        wb = Workbook()
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14)
        bold_font = Font(bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Title
        ws_summary['A1'] = "Fuel Consumption Report"
        ws_summary['A1'].font = title_font
        ws_summary['A2'] = f"Period: {start_date} to {end_date}"
        ws_summary['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        
        # Summary data
        row = 5
        ws_summary[f'A{row}'] = "Total Fuel Records:"
        ws_summary[f'B{row}'] = summary['total_records']
        ws_summary[f'A{row}'].font = bold_font
        
        row += 1
        ws_summary[f'A{row}'] = "Total Litres Consumed:"
        ws_summary[f'B{row}'] = round(summary['total_litres'], 2)
        ws_summary[f'A{row}'].font = bold_font
        
        row += 1
        ws_summary[f'A{row}'] = "Total Cost:"
        ws_summary[f'B{row}'] = round(summary['total_cost'], 2)
        ws_summary[f'A{row}'].font = bold_font
        
        row += 1
        ws_summary[f'A{row}'] = "Average Cost per Litre:"
        ws_summary[f'B{row}'] = round(summary['avg_cost_per_litre'], 2)
        ws_summary[f'A{row}'].font = bold_font
        
        row += 1
        ws_summary[f'A{row}'] = "Average Litres per Fill:"
        ws_summary[f'B{row}'] = round(summary['avg_litres_per_fill'], 2)
        ws_summary[f'A{row}'].font = bold_font
        
        # By Vehicle Sheet
        ws_vehicle = wb.create_sheet("By Vehicle")
        headers = ["Vehicle", "Make/Model", "Fill Count", "Total Litres", "Total Cost", "Distance (km)", "L/100km"]
        
        for col, header in enumerate(headers, 1):
            cell = ws_vehicle.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        row = 2
        for vid, vdata in vehicles_summary.items():
            ws_vehicle.cell(row=row, column=1, value=vdata['vehicle_number'])
            ws_vehicle.cell(row=row, column=2, value=f"{vdata['make']} {vdata['model']}")
            ws_vehicle.cell(row=row, column=3, value=vdata['fill_count'])
            ws_vehicle.cell(row=row, column=4, value=round(vdata['total_litres'], 2))
            ws_vehicle.cell(row=row, column=5, value=round(vdata['total_cost'], 2))
            ws_vehicle.cell(row=row, column=6, value=vdata['total_km'] if vdata['total_km'] > 0 else 'N/A')
            ws_vehicle.cell(row=row, column=7, value=round(vdata['avg_consumption'], 2) if vdata['avg_consumption'] > 0 else 'N/A')
            row += 1
        
        # Detailed Records Sheet
        ws_details = wb.create_sheet("Detailed Records")
        headers = ["Date", "Vehicle", "Driver", "Station", "Odometer", "Distance", "Litres", "L/100km", "Price/L", "Total Cost"]
        
        for col, header in enumerate(headers, 1):
            cell = ws_details.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        row = 2
        for record in fuel_records:
            ws_details.cell(row=row, column=1, value=record['fuel_date'].strftime('%Y-%m-%d'))
            ws_details.cell(row=row, column=2, value=record['vehicle_number'])
            ws_details.cell(row=row, column=3, value=record['driver_name'])
            ws_details.cell(row=row, column=4, value=record['station_name'] if record['station_name'] else 'N/A')
            ws_details.cell(row=row, column=5, value=record['odometer_reading'] if record['odometer_reading'] else 'N/A')
            ws_details.cell(row=row, column=6, value=record['distance_since_last'] if record['distance_since_last'] else '-')
            ws_details.cell(row=row, column=7, value=round(float(record['fuel_amount']), 2))
            ws_details.cell(row=row, column=8, value=round(record['consumption'], 2) if record['consumption'] else '-')
            ws_details.cell(row=row, column=9, value=round(float(record['fuel_cost']) / float(record['fuel_amount']), 2))
            ws_details.cell(row=row, column=10, value=round(float(record['fuel_cost']), 2))
            row += 1
        
        # Auto-adjust column widths
        for ws in [ws_summary, ws_vehicle, ws_details]:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        filename = f"Fuel_Report_{start_date}_to_{end_date}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    finally:
        cursor.close()
        conn.close()

@app.route('/my-fuel-records')
def my_fuel_records():
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Get filter parameter
        vehicle_filter = request.args.get('vehicle_filter', '')
        
        # Build query
        query = """
            SELECT fr.*, 
                   v.vehicle_number, v.make, v.model, v.id as vehicle_id
            FROM fuel_records fr
            JOIN vehicles v ON fr.vehicle_id = v.id
            WHERE fr.employee_id = %s
        """
        params = [session['employee_id']]
        
        if vehicle_filter:
            query += " AND fr.vehicle_id = %s"
            params.append(vehicle_filter)
        
        query += " ORDER BY fr.fuel_date DESC"
        
        cursor.execute(query, params)
        fuel_records = cursor.fetchall()
        
        # Calculate summary statistics
        summary = {
            'total_records': len(fuel_records),
            'total_litres': sum(float(record['fuel_amount']) for record in fuel_records) if fuel_records else 0,
            'total_cost': sum(float(record['fuel_cost']) for record in fuel_records) if fuel_records else 0,
            'avg_cost_per_litre': 0,
            'avg_litres_per_fill': 0,
            'min_cost_per_litre': 0,
            'max_cost_per_litre': 0,
            'min_litres': 0,
            'max_litres': 0,
            'total_distance': 0,
            'avg_consumption': 0,
            'cost_per_km': 0,
            'km_per_litre': 0,
            'records_with_odometer': 0,
            'unique_stations': 0,
            'most_used_station': 'N/A',
            'fuel_type_breakdown': {}
        }
        
        if summary['total_records'] > 0:
            # Basic averages
            summary['avg_litres_per_fill'] = summary['total_litres'] / summary['total_records']
            if summary['total_litres'] > 0:
                summary['avg_cost_per_litre'] = summary['total_cost'] / summary['total_litres']
            
            # Min/Max litres and cost per litre
            litres = [float(record['fuel_amount']) for record in fuel_records]
            costs_per_litre = [float(record['fuel_cost']) / float(record['fuel_amount']) for record in fuel_records]
            summary['min_litres'] = min(litres)
            summary['max_litres'] = max(litres)
            summary['min_cost_per_litre'] = min(costs_per_litre)
            summary['max_cost_per_litre'] = max(costs_per_litre)
            
            # Odometer-based calculations
            records_with_odometer = [r for r in fuel_records if r['odometer_reading']]
            summary['records_with_odometer'] = len(records_with_odometer)
            
            if len(records_with_odometer) >= 2:
                # Sort by odometer reading
                sorted_records = sorted(records_with_odometer, key=lambda x: float(x['odometer_reading']))
                first_odometer = float(sorted_records[0]['odometer_reading'])
                last_odometer = float(sorted_records[-1]['odometer_reading'])
                summary['total_distance'] = last_odometer - first_odometer
                
                if summary['total_distance'] > 0:
                    fuel_for_distance = sum(float(r['fuel_amount']) for r in records_with_odometer)
                    cost_for_distance = sum(float(r['fuel_cost']) for r in records_with_odometer)
                    summary['avg_consumption'] = (fuel_for_distance / summary['total_distance']) * 100  # L/100km
                    summary['km_per_litre'] = summary['total_distance'] / fuel_for_distance
                    summary['cost_per_km'] = cost_for_distance / summary['total_distance']
            
            # Station statistics
            stations = [r['station_name'] for r in fuel_records if r.get('station_name')]
            if stations:
                summary['unique_stations'] = len(set(stations))
                from collections import Counter
                station_counts = Counter(stations)
                summary['most_used_station'] = station_counts.most_common(1)[0][0]
            
            # Fuel type breakdown
            fuel_types = {}
            for record in fuel_records:
                fuel_type = record.get('fuel_type', 'Unknown')
                if fuel_type not in fuel_types:
                    fuel_types[fuel_type] = {'count': 0, 'litres': 0, 'cost': 0}
                fuel_types[fuel_type]['count'] += 1
                fuel_types[fuel_type]['litres'] += float(record['fuel_amount'])
                fuel_types[fuel_type]['cost'] += float(record['fuel_cost'])
            summary['fuel_type_breakdown'] = fuel_types
        
        # Get list of vehicles this employee has fueled
        cursor.execute("""
            SELECT DISTINCT v.id, v.vehicle_number, v.make, v.model
            FROM fuel_records fr
            JOIN vehicles v ON fr.vehicle_id = v.id
            WHERE fr.employee_id = %s
            ORDER BY v.vehicle_number
        """, (session['employee_id'],))
        vehicles = cursor.fetchall()
        
        return render_template('my_fuel_records.html', 
                             fuel_records=fuel_records,
                             summary=summary,
                             vehicles=vehicles,
                             vehicle_filter=vehicle_filter)
    finally:
        cursor.close()
        conn.close()

# Service & Maintenance Routes
@app.route('/service-maintenance')
def service_maintenance():
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Get filter parameters
        vehicle_filter = request.args.get('vehicle_filter', '')
        status_filter = request.args.get('status_filter', '')
        service_type_filter = request.args.get('service_type_filter', '')
        
        # Build query
        query = """
            SELECT sm.*, 
                   v.vehicle_number, v.make, v.model,
                   e.username as performed_by_name
            FROM service_maintenance sm
            JOIN vehicles v ON sm.vehicle_id = v.id
            LEFT JOIN employees e ON sm.performed_by = e.id
            WHERE 1=1
        """
        params = []
        
        if vehicle_filter:
            query += " AND sm.vehicle_id = %s"
            params.append(vehicle_filter)
        
        if status_filter:
            query += " AND sm.status = %s"
            params.append(status_filter)
        
        if service_type_filter:
            query += " AND sm.service_type = %s"
            params.append(service_type_filter)
        
        query += " ORDER BY sm.service_date DESC"
        
        cursor.execute(query, params)
        services = cursor.fetchall()
        
        # Get vehicles and service types for filters
        cursor.execute("SELECT id, vehicle_number, make, model FROM vehicles ORDER BY vehicle_number")
        vehicles = cursor.fetchall()
        
        cursor.execute("SELECT DISTINCT service_type FROM service_maintenance ORDER BY service_type")
        service_types = cursor.fetchall()
        
        return render_template('service_maintenance.html', 
                             services=services,
                             vehicles=vehicles,
                             service_types=service_types,
                             vehicle_filter=vehicle_filter,
                             status_filter=status_filter,
                             service_type_filter=service_type_filter)
    finally:
        cursor.close()
        conn.close()

@app.route('/add-service', methods=['GET', 'POST'])
def add_service():
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    if request.method == 'POST':
        vehicle_id = request.form.get('vehicle_id')
        service_type = request.form.get('service_type')
        service_date = request.form.get('service_date')
        service_provider = request.form.get('service_provider')
        cost = request.form.get('cost') or None
        odometer_reading = request.form.get('odometer_reading') or None
        next_service_date = request.form.get('next_service_date') or None
        next_service_mileage = request.form.get('next_service_mileage') or None
        description = request.form.get('description')
        parts_replaced = request.form.get('parts_replaced')
        status = request.form.get('status', 'completed')
        notes = request.form.get('notes')
        requisition_id = request.form.get('requisition_id') or None
        
        # Handle file upload
        invoice_path = None
        if 'invoice' in request.files:
            file = request.files['invoice']
            if file and file.filename != '':
                if allowed_file(file.filename):
                    from werkzeug.utils import secure_filename
                    import uuid
                    filename = secure_filename(file.filename)
                    unique_filename = f"{uuid.uuid4()}_{filename}"
                    file_path = os.path.join(app.config['UPLOAD_FOLDER_SERVICE'], unique_filename)
                    file.save(file_path)
                    invoice_path = f"uploads/service_invoices/{unique_filename}"
                else:
                    flash('Invalid file type! Allowed types: PNG, JPG, JPEG, PDF', 'danger')
                    return redirect(url_for('add_service'))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("""
                INSERT INTO service_maintenance 
                (vehicle_id, service_type, service_date, service_provider, cost, 
                 odometer_reading, next_service_date, next_service_mileage, description,
                 parts_replaced, invoice_path, status, performed_by, notes)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (vehicle_id, service_type, service_date, service_provider, cost,
                  odometer_reading, next_service_date, next_service_mileage, description,
                  parts_replaced, invoice_path, status, session['employee_id'], notes))
            
            service_id = cursor.lastrowid
            
            # Update vehicle's last_service_date
            cursor.execute("""
                UPDATE vehicles 
                SET last_service_date = %s, mileage = %s
                WHERE id = %s
            """, (service_date, odometer_reading, vehicle_id))
            
            # If created from requisition, link them
            if requisition_id:
                cursor.execute("""
                    UPDATE service_requisitions 
                    SET notes = CONCAT(COALESCE(notes, ''), '\nService Record Created: ID #', %s)
                    WHERE id = %s
                """, (service_id, requisition_id))
            
            conn.commit()
            flash('Service record added successfully!', 'success')
            return redirect(url_for('service_maintenance'))
        finally:
            cursor.close()
            conn.close()
    
    # GET request - show form
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT id, vehicle_number, make, model, mileage, last_service_date
            FROM vehicles 
            ORDER BY vehicle_number
        """)
        vehicles = cursor.fetchall()
        
        # Check if creating from requisition
        requisition_id = request.args.get('requisition_id')
        requisition = None
        
        if requisition_id:
            cursor.execute("""
                SELECT sr.*, 
                       e.username as requester_name,
                       e.email as requester_email,
                       v.vehicle_number as vehicle_reg,
                       v.make as vehicle_make,
                       v.model as vehicle_model
                FROM service_requisitions sr
                JOIN employees e ON sr.requested_by = e.id
                JOIN vehicles v ON sr.vehicle_id = v.id
                WHERE sr.id = %s AND sr.overall_status = 'approved'
            """, (requisition_id,))
            requisition = cursor.fetchone()
            
            if not requisition:
                flash('Requisition not found or not approved!', 'warning')
                requisition_id = None
        
        return render_template('add_service.html', 
                             vehicles=vehicles,
                             requisition=requisition,
                             requisition_id=requisition_id)
    finally:
        cursor.close()
        conn.close()

@app.route('/view-service/<int:service_id>')
def view_service(service_id):
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT sm.*, 
                   v.vehicle_number, v.make, v.model,
                   e.username as performed_by_name
            FROM service_maintenance sm
            JOIN vehicles v ON sm.vehicle_id = v.id
            LEFT JOIN employees e ON sm.performed_by = e.id
            WHERE sm.id = %s
        """, (service_id,))
        service = cursor.fetchone()
        
        if not service:
            flash('Service record not found!', 'danger')
            return redirect(url_for('service_maintenance'))
        
        return render_template('view_service.html', service=service)
    finally:
        cursor.close()
        conn.close()

@app.route('/edit-service/<int:service_id>', methods=['GET', 'POST'])
def edit_service(service_id):
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    if request.method == 'POST':
        service_type = request.form.get('service_type')
        service_date = request.form.get('service_date')
        service_provider = request.form.get('service_provider')
        cost = request.form.get('cost')
        odometer_reading = request.form.get('odometer_reading') or None
        next_service_date = request.form.get('next_service_date') or None
        next_service_mileage = request.form.get('next_service_mileage') or None
        description = request.form.get('description')
        parts_replaced = request.form.get('parts_replaced')
        status = request.form.get('status')
        notes = request.form.get('notes')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            # Get current service record
            cursor.execute("SELECT invoice_path, vehicle_id FROM service_maintenance WHERE id = %s", (service_id,))
            current_service = cursor.fetchone()
            invoice_path = current_service['invoice_path']
            
            # Handle file upload
            if 'invoice' in request.files:
                file = request.files['invoice']
                if file and file.filename != '':
                    if allowed_file(file.filename):
                        # Delete old file if exists
                        if invoice_path:
                            old_file_path = os.path.join('static', invoice_path)
                            if os.path.exists(old_file_path):
                                os.remove(old_file_path)
                        
                        from werkzeug.utils import secure_filename
                        import uuid
                        filename = secure_filename(file.filename)
                        unique_filename = f"{uuid.uuid4()}_{filename}"
                        file_path = os.path.join(app.config['UPLOAD_FOLDER_SERVICE'], unique_filename)
                        file.save(file_path)
                        invoice_path = f"uploads/service_invoices/{unique_filename}"
                    else:
                        flash('Invalid file type! Allowed types: PNG, JPG, JPEG, PDF', 'danger')
                        return redirect(url_for('edit_service', service_id=service_id))
            
            cursor.execute("""
                UPDATE service_maintenance 
                SET service_type = %s, service_date = %s, service_provider = %s, cost = %s,
                    odometer_reading = %s, next_service_date = %s, next_service_mileage = %s,
                    description = %s, parts_replaced = %s, invoice_path = %s, status = %s, notes = %s
                WHERE id = %s
            """, (service_type, service_date, service_provider, cost, odometer_reading,
                  next_service_date, next_service_mileage, description, parts_replaced,
                  invoice_path, status, notes, service_id))
            
            # Update vehicle's last_service_date
            cursor.execute("""
                UPDATE vehicles 
                SET last_service_date = %s, mileage = %s
                WHERE id = %s
            """, (service_date, odometer_reading, current_service['vehicle_id']))
            
            conn.commit()
            flash('Service record updated successfully!', 'success')
            return redirect(url_for('view_service', service_id=service_id))
        finally:
            cursor.close()
            conn.close()
    
    # GET request - show form
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT sm.*, v.vehicle_number, v.make, v.model
            FROM service_maintenance sm
            JOIN vehicles v ON sm.vehicle_id = v.id
            WHERE sm.id = %s
        """, (service_id,))
        service = cursor.fetchone()
        
        if not service:
            flash('Service record not found!', 'danger')
            return redirect(url_for('service_maintenance'))
        
        return render_template('edit_service.html', service=service)
    finally:
        cursor.close()
        conn.close()

@app.route('/delete-service/<int:service_id>')
def delete_service(service_id):
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Get invoice path before deleting
        cursor.execute("SELECT invoice_path FROM service_maintenance WHERE id = %s", (service_id,))
        service = cursor.fetchone()
        
        if service and service['invoice_path']:
            # Delete file from filesystem
            file_path = os.path.join('static', service['invoice_path'])
            if os.path.exists(file_path):
                os.remove(file_path)
        
        cursor.execute("DELETE FROM service_maintenance WHERE id = %s", (service_id,))
        conn.commit()
        flash('Service record deleted successfully!', 'success')
        return redirect(url_for('service_maintenance'))
    finally:
        cursor.close()
        conn.close()

@app.route('/vehicle-service-history/<int:vehicle_id>')
def vehicle_service_history(vehicle_id):
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Get vehicle details
        cursor.execute("""
            SELECT * FROM vehicles WHERE id = %s
        """, (vehicle_id,))
        vehicle = cursor.fetchone()
        
        if not vehicle:
            flash('Vehicle not found!', 'danger')
            return redirect(url_for('vehicles_list'))
        
        # Get service history
        cursor.execute("""
            SELECT sm.*, e.username as performed_by_name
            FROM service_maintenance sm
            LEFT JOIN employees e ON sm.performed_by = e.id
            WHERE sm.vehicle_id = %s
            ORDER BY sm.service_date DESC
        """, (vehicle_id,))
        services = cursor.fetchall()
        
        return render_template('vehicle_service_history.html', vehicle=vehicle, services=services)
    finally:
        cursor.close()
        conn.close()

# Job Card Routes
@app.route('/job-cards')
def job_cards():
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Get filter parameters
        status_filter = request.args.get('status_filter', '')
        vehicle_filter = request.args.get('vehicle_filter', '')
        
        # Build query
        query = """
            SELECT jc.*, 
                   v.vehicle_number, v.make, v.model,
                   e.username as technician_name,
                   creator.username as created_by_name
            FROM job_cards jc
            JOIN vehicles v ON jc.vehicle_id = v.id
            LEFT JOIN employees e ON jc.assigned_technician = e.id
            LEFT JOIN employees creator ON jc.created_by = creator.id
            WHERE 1=1
        """
        params = []
        
        if status_filter:
            query += " AND jc.status = %s"
            params.append(status_filter)
        
        if vehicle_filter:
            query += " AND jc.vehicle_id = %s"
            params.append(vehicle_filter)
        
        query += " ORDER BY jc.created_at DESC"
        
        cursor.execute(query, params)
        job_cards = cursor.fetchall()
        
        # Get vehicles for filter
        cursor.execute("SELECT id, vehicle_number, make, model FROM vehicles ORDER BY vehicle_number")
        vehicles = cursor.fetchall()
        
        return render_template('job_cards.html', 
                             job_cards=job_cards,
                             vehicles=vehicles,
                             status_filter=status_filter,
                             vehicle_filter=vehicle_filter)
    finally:
        cursor.close()
        conn.close()

@app.route('/create-job-card', methods=['GET', 'POST'])
def create_job_card():
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    if request.method == 'POST':
        vehicle_id = request.form.get('vehicle_id')
        customer_name = request.form.get('customer_name')
        customer_phone = request.form.get('customer_phone')
        customer_email = request.form.get('customer_email')
        expected_completion = request.form.get('expected_completion') or None
        odometer_in = request.form.get('odometer_in') or None
        fuel_level = request.form.get('fuel_level')
        reported_issues = request.form.get('reported_issues')
        diagnosis = request.form.get('diagnosis')
        recommended_services = request.form.get('recommended_services')
        assigned_technician = request.form.get('assigned_technician') or None
        priority = request.form.get('priority', 'normal')
        notes = request.form.get('notes')
        requisition_id = request.form.get('requisition_id') or None
        
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            # Generate job card number
            cursor.execute("SELECT COUNT(*) as count FROM job_cards")
            count = cursor.fetchone()['count']
            job_card_number = f"JC{str(count + 1).zfill(6)}"
            
            cursor.execute("""
                INSERT INTO job_cards 
                (job_card_number, vehicle_id, customer_name, customer_phone, customer_email,
                 expected_completion, odometer_in, fuel_level, reported_issues, diagnosis,
                 recommended_services, assigned_technician, priority, notes, created_by, status)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'open')
            """, (job_card_number, vehicle_id, customer_name, customer_phone, customer_email,
                  expected_completion, odometer_in, fuel_level, reported_issues, diagnosis,
                  recommended_services, assigned_technician, priority, notes, session['employee_id']))
            
            job_card_id = cursor.lastrowid
            
            # If created from requisition, update the service record reference
            if requisition_id:
                cursor.execute("""
                    UPDATE service_requisitions 
                    SET notes = CONCAT(COALESCE(notes, ''), '\nJob Card Created: ', %s)
                    WHERE id = %s
                """, (job_card_number, requisition_id))
            
            conn.commit()
            
            flash(f'Job Card {job_card_number} created successfully!', 'success')
            return redirect(url_for('view_job_card', job_card_id=job_card_id))
        finally:
            cursor.close()
            conn.close()
    
    # GET request
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT id, vehicle_number, make, model FROM vehicles ORDER BY vehicle_number")
        vehicles = cursor.fetchall()
        
        cursor.execute("SELECT id, username FROM employees ORDER BY username")
        technicians = cursor.fetchall()
        
        # Check if creating from requisition
        requisition_id = request.args.get('requisition_id')
        requisition = None
        
        if requisition_id:
            cursor.execute("""
                SELECT sr.*, 
                       e.username as requester_name,
                       e.email as requester_email,
                       v.vehicle_number as vehicle_reg,
                       v.make as vehicle_make,
                       v.model as vehicle_model
                FROM service_requisitions sr
                JOIN employees e ON sr.requested_by = e.id
                JOIN vehicles v ON sr.vehicle_id = v.id
                WHERE sr.id = %s AND sr.overall_status = 'approved'
            """, (requisition_id,))
            requisition = cursor.fetchone()
            
            if not requisition:
                flash('Requisition not found or not approved!', 'warning')
                requisition_id = None
        
        return render_template('create_job_card.html', 
                             vehicles=vehicles, 
                             technicians=technicians,
                             requisition=requisition,
                             requisition_id=requisition_id)
    finally:
        cursor.close()
        conn.close()

@app.route('/view-job-card/<int:job_card_id>')
def view_job_card(job_card_id):
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT jc.*, 
                   v.vehicle_number, v.make, v.model, v.year, v.color,
                   e.username as technician_name,
                   creator.username as created_by_name
            FROM job_cards jc
            JOIN vehicles v ON jc.vehicle_id = v.id
            LEFT JOIN employees e ON jc.assigned_technician = e.id
            LEFT JOIN employees creator ON jc.created_by = creator.id
            WHERE jc.id = %s
        """, (job_card_id,))
        job_card = cursor.fetchone()
        
        if not job_card:
            flash('Job card not found!', 'danger')
            return redirect(url_for('job_cards'))
        
        # Get job card items
        cursor.execute("""
            SELECT * FROM job_card_items 
            WHERE job_card_id = %s 
            ORDER BY created_at
        """, (job_card_id,))
        items = cursor.fetchall()
        
        # Get related service record if job card is completed
        service_record = None
        if job_card['status'] == 'completed':
            cursor.execute("""
                SELECT id FROM service_maintenance 
                WHERE job_card_id = %s
            """, (job_card_id,))
            service_record = cursor.fetchone()
        
        return render_template('view_job_card.html', job_card=job_card, items=items, service_record=service_record)
    finally:
        cursor.close()
        conn.close()

@app.route('/edit-job-card/<int:job_card_id>', methods=['GET', 'POST'])
def edit_job_card(job_card_id):
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    if request.method == 'POST':
        customer_name = request.form.get('customer_name')
        customer_phone = request.form.get('customer_phone')
        customer_email = request.form.get('customer_email')
        expected_completion = request.form.get('expected_completion')
        odometer_in = request.form.get('odometer_in') or None
        odometer_out = request.form.get('odometer_out') or None
        fuel_level = request.form.get('fuel_level')
        reported_issues = request.form.get('reported_issues')
        diagnosis = request.form.get('diagnosis')
        recommended_services = request.form.get('recommended_services')
        assigned_technician = request.form.get('assigned_technician') or None
        priority = request.form.get('priority')
        status = request.form.get('status')
        labor_cost = request.form.get('labor_cost') or 0
        notes = request.form.get('notes')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            # Get current job card data
            cursor.execute("SELECT status, vehicle_id FROM job_cards WHERE id = %s", (job_card_id,))
            old_job_card = cursor.fetchone()
            old_status = old_job_card['status']
            vehicle_id = old_job_card['vehicle_id']
            
            # Calculate parts cost from items
            cursor.execute("""
                SELECT SUM(total_price) as parts_total 
                FROM job_card_items 
                WHERE job_card_id = %s AND item_type = 'part'
            """, (job_card_id,))
            result = cursor.fetchone()
            parts_cost = result['parts_total'] if result['parts_total'] else 0
            
            total_cost = float(labor_cost) + float(parts_cost)
            
            cursor.execute("""
                UPDATE job_cards 
                SET customer_name = %s, customer_phone = %s, customer_email = %s,
                    expected_completion = %s, odometer_in = %s, odometer_out = %s,
                    fuel_level = %s, reported_issues = %s, diagnosis = %s,
                    recommended_services = %s, assigned_technician = %s, priority = %s,
                    status = %s, labor_cost = %s, parts_cost = %s, total_cost = %s, notes = %s
                WHERE id = %s
            """, (customer_name, customer_phone, customer_email, expected_completion,
                  odometer_in, odometer_out, fuel_level, reported_issues, diagnosis,
                  recommended_services, assigned_technician, priority, status,
                  labor_cost, parts_cost, total_cost, notes, job_card_id))
            
            # Update date_out if status changed to completed
            if status == 'completed':
                cursor.execute("""
                    UPDATE job_cards SET date_out = CURRENT_TIMESTAMP 
                    WHERE id = %s AND date_out IS NULL
                """, (job_card_id,))
                
                # Create service record if status changed from non-completed to completed
                if old_status != 'completed':
                    # Get job card details
                    cursor.execute("""
                        SELECT jc.*, v.vehicle_number 
                        FROM job_cards jc
                        JOIN vehicles v ON jc.vehicle_id = v.id
                        WHERE jc.id = %s
                    """, (job_card_id,))
                    jc = cursor.fetchone()
                    
                    # Get parts list for service record
                    cursor.execute("""
                        SELECT description, quantity FROM job_card_items 
                        WHERE job_card_id = %s AND item_type = 'part'
                    """, (job_card_id,))
                    parts = cursor.fetchall()
                    parts_list = "\n".join([f"- {p['description']} (Qty: {p['quantity']})" for p in parts]) if parts else None
                    
                    # Determine service type from diagnosis or recommended services
                    service_type = "General Service"
                    if diagnosis:
                        if any(word in diagnosis.lower() for word in ['oil', 'change']):
                            service_type = "Oil Change"
                        elif any(word in diagnosis.lower() for word in ['brake', 'brakes']):
                            service_type = "Brake Service"
                        elif any(word in diagnosis.lower() for word in ['tire', 'tires']):
                            service_type = "Tire Rotation"
                        elif any(word in diagnosis.lower() for word in ['engine', 'tune']):
                            service_type = "Engine Tune-up"
                        elif any(word in diagnosis.lower() for word in ['inspection']):
                            service_type = "Inspection"
                        elif any(word in diagnosis.lower() for word in ['repair']):
                            service_type = "Repair"
                    
                    # Calculate next service mileage (current mileage + 10000 km)
                    next_service_mileage = int(odometer_out) + 10000 if odometer_out else None
                    
                    # Create service maintenance record
                    cursor.execute("""
                        INSERT INTO service_maintenance 
                        (vehicle_id, service_type, service_date, cost, odometer_reading,
                         next_service_mileage, description, parts_replaced, status, performed_by, job_card_id, notes)
                        VALUES (%s, %s, CURDATE(), %s, %s, %s, %s, %s, 'completed', %s, %s, %s)
                    """, (vehicle_id, service_type, total_cost, odometer_out, next_service_mileage,
                          diagnosis if diagnosis else reported_issues, parts_list,
                          assigned_technician if assigned_technician else session['employee_id'],
                          job_card_id,
                          f"Auto-generated from Job Card: {jc['job_card_number']}\n{notes if notes else ''}"))
                    
                    # Update vehicle's last_service_date and mileage
                    cursor.execute("""
                        UPDATE vehicles 
                        SET last_service_date = CURDATE(), mileage = %s
                        WHERE id = %s
                    """, (odometer_out, vehicle_id))
                    
                    flash('Job card completed and service record created!', 'success')
                    
                    # Send WhatsApp notification for job card completion
                    try:
                        cursor.execute("""
                            SELECT ns.job_card_status_whatsapp, ns.whatsapp_number
                            FROM notification_settings ns
                            WHERE ns.employee_id = %s AND ns.job_card_status_whatsapp = TRUE
                        """, (session['employee_id'],))
                        notif_pref = cursor.fetchone()
                        if notif_pref and notif_pref['whatsapp_number']:
                            whatsapp_service.send_job_card_status_alert(
                                jc['job_card_number'], 
                                'completed', 
                                jc['vehicle_number'],
                                notif_pref['whatsapp_number']
                            )
                    except Exception as e:
                        print(f"WhatsApp notification error: {e}")
            
            conn.commit()
            if status != 'completed' or old_status == 'completed':
                flash('Job card updated successfully!', 'success')
            
            return redirect(url_for('view_job_card', job_card_id=job_card_id))
        finally:
            cursor.close()
            conn.close()
    
    # GET request
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT jc.*, v.vehicle_number, v.make, v.model
            FROM job_cards jc
            JOIN vehicles v ON jc.vehicle_id = v.id
            WHERE jc.id = %s
        """, (job_card_id,))
        job_card = cursor.fetchone()
        
        if not job_card:
            flash('Job card not found!', 'danger')
            return redirect(url_for('job_cards'))
        
        cursor.execute("SELECT id, username FROM employees ORDER BY username")
        technicians = cursor.fetchall()
        
        return render_template('edit_job_card.html', job_card=job_card, technicians=technicians)
    finally:
        cursor.close()
        conn.close()

@app.route('/add-job-card-item/<int:job_card_id>', methods=['POST'])
def add_job_card_item(job_card_id):
    if 'employee_id' not in session:
        return redirect(url_for('employee_login'))
    
    item_type = request.form.get('item_type')
    description = request.form.get('description')
    quantity = request.form.get('quantity', 1)
    unit_price = request.form.get('unit_price', 0)
    
    total_price = float(quantity) * float(unit_price)
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO job_card_items 
            (job_card_id, item_type, description, quantity, unit_price, total_price)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (job_card_id, item_type, description, quantity, unit_price, total_price))
        
        # Update job card costs
        cursor.execute("""
            SELECT 
                SUM(CASE WHEN item_type = 'part' THEN total_price ELSE 0 END) as parts_total,
                SUM(CASE WHEN item_type = 'labor' THEN total_price ELSE 0 END) as labor_total
            FROM job_card_items 
            WHERE job_card_id = %s
        """, (job_card_id,))
        result = cursor.fetchone()
        
        parts_cost = result['parts_total'] if result['parts_total'] else 0
        labor_cost = result['labor_total'] if result['labor_total'] else 0
        total_cost = parts_cost + labor_cost
        
        cursor.execute("""
            UPDATE job_cards 
            SET parts_cost = %s, labor_cost = %s, total_cost = %s
            WHERE id = %s
        """, (parts_cost, labor_cost, total_cost, job_card_id))
        
        conn.commit()
        flash('Item added successfully!', 'success')
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('view_job_card', job_card_id=job_card_id))

@app.route('/delete-job-card-item/<int:item_id>/<int:job_card_id>')
def delete_job_card_item(item_id, job_card_id):
    if 'employee_id' not in session:
        return redirect(url_for('employee_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM job_card_items WHERE id = %s", (item_id,))
        
        # Update job card costs
        cursor.execute("""
            SELECT 
                SUM(CASE WHEN item_type = 'part' THEN total_price ELSE 0 END) as parts_total,
                SUM(CASE WHEN item_type = 'labor' THEN total_price ELSE 0 END) as labor_total
            FROM job_card_items 
            WHERE job_card_id = %s
        """, (job_card_id,))
        result = cursor.fetchone()
        
        parts_cost = result['parts_total'] if result['parts_total'] else 0
        labor_cost = result['labor_total'] if result['labor_total'] else 0
        total_cost = parts_cost + labor_cost
        
        cursor.execute("""
            UPDATE job_cards 
            SET parts_cost = %s, labor_cost = %s, total_cost = %s
            WHERE id = %s
        """, (parts_cost, labor_cost, total_cost, job_card_id))
        
        conn.commit()
        flash('Item deleted successfully!', 'success')
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('view_job_card', job_card_id=job_card_id))

@app.route('/print-job-card/<int:job_card_id>')
def print_job_card(job_card_id):
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT jc.*, 
                   v.vehicle_number, v.make, v.model, v.year, v.color,
                   e.username as technician_name,
                   creator.username as created_by_name
            FROM job_cards jc
            JOIN vehicles v ON jc.vehicle_id = v.id
            LEFT JOIN employees e ON jc.assigned_technician = e.id
            LEFT JOIN employees creator ON jc.created_by = creator.id
            WHERE jc.id = %s
        """, (job_card_id,))
        job_card = cursor.fetchone()
        
        if not job_card:
            flash('Job card not found!', 'danger')
            return redirect(url_for('job_cards'))
        
        cursor.execute("""
            SELECT * FROM job_card_items 
            WHERE job_card_id = %s 
            ORDER BY created_at
        """, (job_card_id,))
        items = cursor.fetchall()
        
        # Get print settings
        cursor.execute("SELECT * FROM print_settings LIMIT 1")
        print_config = cursor.fetchone()
        
        return render_template('print_job_card.html', job_card=job_card, items=items, print_config=print_config)
    finally:
        cursor.close()
        conn.close()

@app.route('/delete-job-card/<int:job_card_id>')
def delete_job_card(job_card_id):
    if 'employee_id' not in session:
        return redirect(url_for('employee_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM job_cards WHERE id = %s", (job_card_id,))
        conn.commit()
        flash('Job card deleted successfully!', 'success')
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('job_cards'))

@app.route('/service-notifications')
def service_notifications():
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Get vehicles with service due based on latest fuel tracking mileage
        cursor.execute("""
            SELECT 
                v.id,
                v.vehicle_number,
                v.make,
                v.model,
                v.year,
                v.mileage as vehicle_mileage,
                COALESCE(fr.latest_mileage, v.mileage) as latest_mileage,
                sm.next_service_mileage,
                sm.next_service_date,
                sm.service_type as last_service_type,
                sm.service_date as last_service_date,
                v.last_service_date as vehicle_last_service,
                CASE 
                    WHEN sm.next_service_mileage IS NOT NULL AND COALESCE(fr.latest_mileage, v.mileage) IS NOT NULL 
                    THEN sm.next_service_mileage - COALESCE(fr.latest_mileage, v.mileage)
                    ELSE NULL
                END as km_until_service,
                CASE
                    WHEN sm.next_service_date IS NOT NULL
                    THEN DATEDIFF(sm.next_service_date, CURDATE())
                    ELSE NULL
                END as days_until_service
            FROM vehicles v
            LEFT JOIN (
                SELECT vehicle_id, MAX(odometer_reading) as latest_mileage
                FROM fuel_records
                WHERE odometer_reading IS NOT NULL
                GROUP BY vehicle_id
            ) fr ON v.id = fr.vehicle_id
            LEFT JOIN (
                SELECT sm1.vehicle_id, sm1.next_service_mileage, sm1.next_service_date, 
                       sm1.service_type, sm1.service_date
                FROM service_maintenance sm1
                INNER JOIN (
                    SELECT vehicle_id, MAX(service_date) as max_date
                    FROM service_maintenance
                    WHERE next_service_mileage IS NOT NULL OR next_service_date IS NOT NULL
                    GROUP BY vehicle_id
                ) sm2 ON sm1.vehicle_id = sm2.vehicle_id AND sm1.service_date = sm2.max_date
            ) sm ON v.id = sm.vehicle_id
            ORDER BY 
                CASE 
                    WHEN sm.next_service_mileage IS NOT NULL AND COALESCE(fr.latest_mileage, v.mileage) IS NOT NULL 
                    THEN sm.next_service_mileage - COALESCE(fr.latest_mileage, v.mileage)
                    ELSE 999999
                END ASC,
                v.vehicle_number
        """)
        
        vehicles = cursor.fetchall()
        
        # Categorize vehicles
        overdue = []
        due_soon = []
        upcoming = []
        no_data = []
        
        for vehicle in vehicles:
            # Skip if both service indicators are None
            if vehicle['km_until_service'] is None and vehicle['days_until_service'] is None:
                no_data.append(vehicle)
            # Check if overdue (negative or zero remaining)
            elif (vehicle['km_until_service'] is not None and vehicle['km_until_service'] <= 0) or \
                 (vehicle['days_until_service'] is not None and vehicle['days_until_service'] <= 0):
                overdue.append(vehicle)
            # Check if due soon (within 1000 km or 7 days, but greater than 0)
            elif (vehicle['km_until_service'] is not None and 0 < vehicle['km_until_service'] <= 1000) or \
                 (vehicle['days_until_service'] is not None and 0 < vehicle['days_until_service'] <= 7):
                due_soon.append(vehicle)
            # Otherwise it's upcoming
            elif vehicle['km_until_service'] is not None or vehicle['days_until_service'] is not None:
                upcoming.append(vehicle)
            else:
                no_data.append(vehicle)
        
        # Send WhatsApp notifications for overdue and due soon services
        try:
            cursor.execute("""
                SELECT ns.service_overdue_whatsapp, ns.service_due_soon_whatsapp, 
                       ns.whatsapp_number, ns.employee_id
                FROM notification_settings ns
                WHERE (ns.service_overdue_whatsapp = TRUE OR ns.service_due_soon_whatsapp = TRUE)
                  AND ns.whatsapp_number IS NOT NULL
            """)
            notif_prefs = cursor.fetchall()
            
            for pref in notif_prefs:
                # Send overdue alerts
                if pref['service_overdue_whatsapp']:
                    for vehicle in overdue[:3]:  # Limit to 3 alerts per check
                        vehicle_details = {
                            'make': vehicle['make'],
                            'model': vehicle['model'],
                            'current_mileage': vehicle['latest_mileage'],
                            'service_due_mileage': vehicle['next_service_mileage'],
                            'overdue_km': abs(vehicle['km_until_service']) if vehicle['km_until_service'] else 'N/A'
                        }
                        whatsapp_service.send_service_overdue_alert(
                            vehicle['vehicle_number'],
                            vehicle_details,
                            pref['whatsapp_number']
                        )
                
                # Send due soon alerts
                if pref['service_due_soon_whatsapp']:
                    for vehicle in due_soon[:3]:  # Limit to 3 alerts per check
                        vehicle_details = {
                            'make': vehicle['make'],
                            'model': vehicle['model'],
                            'current_mileage': vehicle['latest_mileage'],
                            'service_due_mileage': vehicle['next_service_mileage'],
                            'km_remaining': vehicle['km_until_service'] if vehicle['km_until_service'] else 'N/A'
                        }
                        whatsapp_service.send_service_due_soon_alert(
                            vehicle['vehicle_number'],
                            vehicle_details,
                            pref['whatsapp_number']
                        )
        except Exception as e:
            print(f"WhatsApp notification error: {e}")
        
        return render_template('service_notifications.html', 
                             overdue=overdue, 
                             due_soon=due_soon, 
                             upcoming=upcoming,
                             no_data=no_data)
    finally:
        cursor.close()
        conn.close()

@app.route('/service-notifications/print')
def service_notifications_print():
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Same query as service_notifications
        cursor.execute("""
            SELECT 
                v.id,
                v.vehicle_number,
                v.make,
                v.model,
                v.year,
                v.mileage as vehicle_mileage,
                COALESCE(fr.latest_mileage, v.mileage) as latest_mileage,
                sm.next_service_mileage,
                sm.next_service_date,
                sm.service_type as last_service_type,
                sm.service_date as last_service_date,
                v.last_service_date as vehicle_last_service,
                CASE 
                    WHEN sm.next_service_mileage IS NOT NULL AND COALESCE(fr.latest_mileage, v.mileage) IS NOT NULL 
                    THEN sm.next_service_mileage - COALESCE(fr.latest_mileage, v.mileage)
                    ELSE NULL
                END as km_until_service,
                CASE
                    WHEN sm.next_service_date IS NOT NULL
                    THEN DATEDIFF(sm.next_service_date, CURDATE())
                    ELSE NULL
                END as days_until_service
            FROM vehicles v
            LEFT JOIN (
                SELECT vehicle_id, MAX(odometer_reading) as latest_mileage
                FROM fuel_records
                WHERE odometer_reading IS NOT NULL
                GROUP BY vehicle_id
            ) fr ON v.id = fr.vehicle_id
            LEFT JOIN (
                SELECT sm1.vehicle_id, sm1.next_service_mileage, sm1.next_service_date, 
                       sm1.service_type, sm1.service_date
                FROM service_maintenance sm1
                INNER JOIN (
                    SELECT vehicle_id, MAX(service_date) as max_date
                    FROM service_maintenance
                    WHERE next_service_mileage IS NOT NULL OR next_service_date IS NOT NULL
                    GROUP BY vehicle_id
                ) sm2 ON sm1.vehicle_id = sm2.vehicle_id AND sm1.service_date = sm2.max_date
            ) sm ON v.id = sm.vehicle_id
            ORDER BY 
                CASE 
                    WHEN sm.next_service_mileage IS NOT NULL AND COALESCE(fr.latest_mileage, v.mileage) IS NOT NULL 
                    THEN sm.next_service_mileage - COALESCE(fr.latest_mileage, v.mileage)
                    ELSE 999999
                END ASC,
                v.vehicle_number
        """)
        
        vehicles = cursor.fetchall()
        
        # Categorize vehicles
        overdue = []
        due_soon = []
        upcoming = []
        no_data = []
        
        for vehicle in vehicles:
            if vehicle['km_until_service'] is None and vehicle['days_until_service'] is None:
                no_data.append(vehicle)
            elif (vehicle['km_until_service'] is not None and vehicle['km_until_service'] <= 0) or \
                 (vehicle['days_until_service'] is not None and vehicle['days_until_service'] <= 0):
                overdue.append(vehicle)
            elif (vehicle['km_until_service'] is not None and 0 < vehicle['km_until_service'] <= 1000) or \
                 (vehicle['days_until_service'] is not None and 0 < vehicle['days_until_service'] <= 7):
                due_soon.append(vehicle)
            elif vehicle['km_until_service'] is not None or vehicle['days_until_service'] is not None:
                upcoming.append(vehicle)
            else:
                no_data.append(vehicle)
        
        # Get print settings
        cursor.execute("SELECT * FROM print_settings LIMIT 1")
        print_config = cursor.fetchone()
        
        return render_template('service_notifications_print.html', 
                             overdue=overdue, 
                             due_soon=due_soon, 
                             upcoming=upcoming,
                             no_data=no_data,
                             print_config=print_config)
    finally:
        cursor.close()
        conn.close()

# Settings Routes
@app.route('/settings')
def settings():
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Get system settings grouped by category
        cursor.execute("""
            SELECT * FROM system_settings 
            ORDER BY category, setting_key
        """)
        all_settings = cursor.fetchall()
        
        # Group settings by category
        settings_by_category = {}
        for setting in all_settings:
            category = setting['category'] or 'other'
            if category not in settings_by_category:
                settings_by_category[category] = []
            settings_by_category[category].append(setting)
        
        return render_template('settings.html', settings_by_category=settings_by_category)
    finally:
        cursor.close()
        conn.close()

@app.route('/settings/update', methods=['POST'])
def update_settings():
    if 'employee_id' not in session:
        return redirect(url_for('employee_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        for key, value in request.form.items():
            if key.startswith('setting_'):
                setting_key = key.replace('setting_', '')
                cursor.execute("""
                    UPDATE system_settings 
                    SET setting_value = %s, updated_by = %s 
                    WHERE setting_key = %s AND is_editable = TRUE
                """, (value, session['employee_id'], setting_key))
        
        conn.commit()
        flash('Settings updated successfully!', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Error updating settings: {str(e)}', 'danger')
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('settings'))

@app.route('/settings/profile', methods=['GET', 'POST'])
def profile_settings():
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        department = request.form.get('department')
        position = request.form.get('position')
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        
        try:
            # Update profile information
            cursor.execute("""
                UPDATE employees 
                SET username = %s, email = %s, department = %s, position = %s
                WHERE id = %s
            """, (username, email, department, position, session['employee_id']))
            
            # Update password if provided
            if current_password and new_password:
                cursor.execute("SELECT password_hash FROM employees WHERE id = %s", (session['employee_id'],))
                employee = cursor.fetchone()
                
                if check_password_hash(employee['password_hash'], current_password):
                    if new_password == confirm_password:
                        hashed_password = generate_password_hash(new_password)
                        cursor.execute("""
                            UPDATE employees SET password_hash = %s WHERE id = %s
                        """, (hashed_password, session['employee_id']))
                        flash('Password updated successfully!', 'success')
                    else:
                        flash('New passwords do not match!', 'danger')
                        conn.rollback()
                        cursor.close()
                        conn.close()
                        return redirect(url_for('profile_settings'))
                else:
                    flash('Current password is incorrect!', 'danger')
                    conn.rollback()
                    cursor.close()
                    conn.close()
                    return redirect(url_for('profile_settings'))
            
            conn.commit()
            flash('Profile updated successfully!', 'success')
            return redirect(url_for('profile_settings'))
        except Exception as e:
            conn.rollback()
            flash(f'Error updating profile: {str(e)}', 'danger')
    
    try:
        cursor.execute("SELECT * FROM employees WHERE id = %s", (session['employee_id'],))
        employee = cursor.fetchone()
        return render_template('profile_settings.html', employee=employee)
    finally:
        cursor.close()
        conn.close()

@app.route('/settings/notifications', methods=['GET', 'POST'])
def notification_settings():
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    if request.method == 'POST':
        try:
            # Email notifications
            service_overdue_email = 1 if request.form.get('service_overdue_email') else 0
            service_due_soon_email = 1 if request.form.get('service_due_soon_email') else 0
            fuel_expense_alert = 1 if request.form.get('fuel_expense_alert') else 0
            vehicle_assignment_email = 1 if request.form.get('vehicle_assignment_email') else 0
            job_card_status_email = 1 if request.form.get('job_card_status_email') else 0
            daily_summary_email = 1 if request.form.get('daily_summary_email') else 0
            
            # WhatsApp notifications
            service_overdue_whatsapp = 1 if request.form.get('service_overdue_whatsapp') else 0
            service_due_soon_whatsapp = 1 if request.form.get('service_due_soon_whatsapp') else 0
            fuel_expense_alert_whatsapp = 1 if request.form.get('fuel_expense_alert_whatsapp') else 0
            vehicle_assignment_whatsapp = 1 if request.form.get('vehicle_assignment_whatsapp') else 0
            job_card_status_whatsapp = 1 if request.form.get('job_card_status_whatsapp') else 0
            daily_summary_whatsapp = 1 if request.form.get('daily_summary_whatsapp') else 0
            whatsapp_number = request.form.get('whatsapp_number', '')
            
            cursor.execute("""
                INSERT INTO notification_settings 
                (employee_id, service_overdue_email, service_due_soon_email, fuel_expense_alert,
                 vehicle_assignment_email, job_card_status_email, daily_summary_email,
                 service_overdue_whatsapp, service_due_soon_whatsapp, fuel_expense_alert_whatsapp,
                 vehicle_assignment_whatsapp, job_card_status_whatsapp, daily_summary_whatsapp, whatsapp_number)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    service_overdue_email = VALUES(service_overdue_email),
                    service_due_soon_email = VALUES(service_due_soon_email),
                    fuel_expense_alert = VALUES(fuel_expense_alert),
                    vehicle_assignment_email = VALUES(vehicle_assignment_email),
                    job_card_status_email = VALUES(job_card_status_email),
                    daily_summary_email = VALUES(daily_summary_email),
                    service_overdue_whatsapp = VALUES(service_overdue_whatsapp),
                    service_due_soon_whatsapp = VALUES(service_due_soon_whatsapp),
                    fuel_expense_alert_whatsapp = VALUES(fuel_expense_alert_whatsapp),
                    vehicle_assignment_whatsapp = VALUES(vehicle_assignment_whatsapp),
                    job_card_status_whatsapp = VALUES(job_card_status_whatsapp),
                    daily_summary_whatsapp = VALUES(daily_summary_whatsapp),
                    whatsapp_number = VALUES(whatsapp_number)
            """, (session['employee_id'], service_overdue_email, service_due_soon_email,
                  fuel_expense_alert, vehicle_assignment_email, job_card_status_email, daily_summary_email,
                  service_overdue_whatsapp, service_due_soon_whatsapp, fuel_expense_alert_whatsapp,
                  vehicle_assignment_whatsapp, job_card_status_whatsapp, daily_summary_whatsapp, whatsapp_number))
            
            conn.commit()
            flash('Notification settings updated successfully!', 'success')
            return redirect(url_for('notification_settings'))
        except Exception as e:
            conn.rollback()
            flash(f'Error updating notification settings: {str(e)}', 'danger')
    
    try:
        cursor.execute("""
            SELECT * FROM notification_settings WHERE employee_id = %s
        """, (session['employee_id'],))
        settings = cursor.fetchone()
        
        # Create default settings if none exist
        if not settings:
            cursor.execute("""
                INSERT INTO notification_settings (employee_id) VALUES (%s)
            """, (session['employee_id'],))
            conn.commit()
            cursor.execute("""
                SELECT * FROM notification_settings WHERE employee_id = %s
            """, (session['employee_id'],))
            settings = cursor.fetchone()
        
        return render_template('notification_settings.html', settings=settings)
    finally:
        cursor.close()
        conn.close()

@app.route('/settings/print')
def print_settings():
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT * FROM print_settings LIMIT 1")
        print_config = cursor.fetchone()
        return render_template('print_settings.html', print_config=print_config)
    finally:
        cursor.close()
        conn.close()

@app.route('/settings/print/update', methods=['POST'])
def update_print_settings():
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    company_name = request.form.get('company_name')
    company_tagline = request.form.get('company_tagline')
    company_address = request.form.get('company_address')
    company_phone = request.form.get('company_phone')
    company_email = request.form.get('company_email')
    company_website = request.form.get('company_website')
    footer_left = request.form.get('footer_left')
    footer_center = request.form.get('footer_center')
    footer_right = request.form.get('footer_right')
    remove_logo = request.form.get('remove_logo')
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Handle logo upload
        logo_path = None
        if 'logo' in request.files:
            file = request.files['logo']
            if file and file.filename != '':
                if allowed_file(file.filename):
                    from werkzeug.utils import secure_filename
                    import uuid
                    
                    # Create logos directory if it doesn't exist
                    logo_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'logos')
                    os.makedirs(logo_dir, exist_ok=True)
                    
                    filename = secure_filename(file.filename)
                    unique_filename = f"{uuid.uuid4()}_{filename}"
                    file_path = os.path.join(logo_dir, unique_filename)
                    file.save(file_path)
                    logo_path = f"uploads/logos/{unique_filename}"
        
        # Check if settings exist
        cursor.execute("SELECT id, logo_path FROM print_settings LIMIT 1")
        existing = cursor.fetchone()
        
        if existing:
            # Update existing settings
            if remove_logo:
                # Delete old logo file if exists
                if existing['logo_path']:
                    old_logo_path = os.path.join(app.config['UPLOAD_FOLDER'], '..', 'static', existing['logo_path'])
                    if os.path.exists(old_logo_path):
                        os.remove(old_logo_path)
                
                cursor.execute("""
                    UPDATE print_settings SET
                        company_name = %s,
                        company_tagline = %s,
                        company_address = %s,
                        company_phone = %s,
                        company_email = %s,
                        company_website = %s,
                        logo_path = NULL,
                        footer_left = %s,
                        footer_center = %s,
                        footer_right = %s,
                        updated_by = %s
                    WHERE id = %s
                """, (company_name, company_tagline, company_address, company_phone, company_email,
                      company_website, footer_left, footer_center, footer_right,
                      session['employee_id'], existing['id']))
            elif logo_path:
                # Delete old logo if new one is uploaded
                if existing['logo_path']:
                    old_logo_path = os.path.join(app.config['UPLOAD_FOLDER'], '..', 'static', existing['logo_path'])
                    if os.path.exists(old_logo_path):
                        os.remove(old_logo_path)
                
                cursor.execute("""
                    UPDATE print_settings SET
                        company_name = %s,
                        company_tagline = %s,
                        company_address = %s,
                        company_phone = %s,
                        company_email = %s,
                        company_website = %s,
                        logo_path = %s,
                        footer_left = %s,
                        footer_center = %s,
                        footer_right = %s,
                        updated_by = %s
                    WHERE id = %s
                """, (company_name, company_tagline, company_address, company_phone, company_email,
                      company_website, logo_path, footer_left, footer_center, footer_right,
                      session['employee_id'], existing['id']))
            else:
                cursor.execute("""
                    UPDATE print_settings SET
                        company_name = %s,
                        company_tagline = %s,
                        company_address = %s,
                        company_phone = %s,
                        company_email = %s,
                        company_website = %s,
                        footer_left = %s,
                        footer_center = %s,
                        footer_right = %s,
                        updated_by = %s
                    WHERE id = %s
                """, (company_name, company_tagline, company_address, company_phone, company_email,
                      company_website, footer_left, footer_center, footer_right,
                      session['employee_id'], existing['id']))
        else:
            # Insert new settings
            cursor.execute("""
                INSERT INTO print_settings 
                (company_name, company_tagline, company_address, company_phone, company_email,
                 company_website, logo_path, footer_left, footer_center, footer_right, updated_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (company_name, company_tagline, company_address, company_phone, company_email,
                  company_website, logo_path, footer_left, footer_center, footer_right,
                  session['employee_id']))
        
        conn.commit()
        flash('Print settings updated successfully!', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Error updating print settings: {str(e)}', 'danger')
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('print_settings'))

@app.route('/settings/backups')
def backup_settings():
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    from backup_manager import list_backups, get_backup_size_total
    
    backups = list_backups()
    total_size = get_backup_size_total()
    
    return render_template('backup_settings.html', backups=backups, total_size=total_size)

@app.route('/settings/backups/create', methods=['POST'])
def create_backup_route():
    if 'employee_id' not in session:
        return redirect(url_for('employee_login'))
    
    from backup_manager import create_backup
    
    try:
        backup_path = create_backup()
        flash(f'Backup created successfully: {os.path.basename(backup_path)}', 'success')
    except Exception as e:
        flash(f'Error creating backup: {str(e)}', 'danger')
    
    return redirect(url_for('backup_settings'))

@app.route('/settings/backups/delete/<filename>', methods=['POST'])
def delete_backup_route(filename):
    if 'employee_id' not in session:
        return redirect(url_for('employee_login'))
    
    from backup_manager import delete_backup
    
    if delete_backup(filename):
        flash('Backup deleted successfully!', 'success')
    else:
        flash('Backup file not found!', 'danger')
    
    return redirect(url_for('backup_settings'))

@app.route('/settings/backups/download/<filename>')
def download_backup(filename):
    if 'employee_id' not in session:
        return redirect(url_for('employee_login'))
    
    from backup_manager import BACKUP_DIR
    from flask import send_from_directory
    
    return send_from_directory(BACKUP_DIR, filename, as_attachment=True)

@app.route('/settings/backups/restore/<filename>', methods=['POST'])
def restore_backup_route(filename):
    if 'employee_id' not in session:
        return redirect(url_for('employee_login'))
    
    from backup_manager import restore_backup
    
    try:
        restore_backup(filename)
        flash('Database restored successfully! Please log in again.', 'success')
        session.clear()
        return redirect(url_for('employee_login'))
    except Exception as e:
        flash(f'Error restoring backup: {str(e)}', 'danger')
        return redirect(url_for('backup_settings'))

# Test WhatsApp notification route
@app.route('/test-whatsapp', methods=['GET', 'POST'])
def test_whatsapp():
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    if request.method == 'POST':
        phone_number = request.form.get('phone_number')
        message_type = request.form.get('message_type')
        
        if not phone_number:
            flash('Phone number is required!', 'danger')
            return redirect(url_for('test_whatsapp'))
        
        try:
            # Format phone number
            if not phone_number.startswith('+'):
                phone_number = '+' + phone_number
            
            # Send test message based on type
            if message_type == 'simple':
                result = whatsapp_service.send_whatsapp_message(
                    phone_number,
                    "Test message from Fleet Management System. WhatsApp notifications are working!"
                )
            elif message_type == 'service_overdue':
                vehicle_details = {
                    'make': 'Toyota',
                    'model': 'Hilux',
                    'current_mileage': 105000,
                    'service_due_mileage': 100000,
                    'overdue_km': 5000
                }
                result = whatsapp_service.send_service_overdue_alert(
                    'TEST-001',
                    vehicle_details,
                    phone_number
                )
            elif message_type == 'service_due_soon':
                vehicle_details = {
                    'make': 'Toyota',
                    'model': 'Hilux',
                    'current_mileage': 99500,
                    'service_due_mileage': 100000,
                    'km_remaining': 500
                }
                result = whatsapp_service.send_service_due_soon_alert(
                    'TEST-001',
                    vehicle_details,
                    phone_number
                )
            elif message_type == 'vehicle_assignment':
                vehicle_details = {
                    'make': 'Toyota',
                    'model': 'Hilux',
                    'assignment_date': 'Today',
                    'purpose': 'Field Work'
                }
                result = whatsapp_service.send_vehicle_assignment_alert(
                    'TEST-001',
                    vehicle_details,
                    'Test User',
                    phone_number
                )
            elif message_type == 'job_card_status':
                result = whatsapp_service.send_job_card_status_alert(
                    'JC-TEST-001',
                    'completed',
                    'TEST-001',
                    phone_number
                )
            else:
                flash('Invalid message type!', 'danger')
                return redirect(url_for('test_whatsapp'))
            
            if result:
                flash('WhatsApp message sent successfully!', 'success')
            else:
                flash('Failed to send WhatsApp message. Check settings and logs.', 'danger')
                
        except Exception as e:
            flash(f'Error sending WhatsApp: {str(e)}', 'danger')
        
        return redirect(url_for('test_whatsapp'))
    
    # GET request
    return render_template('test_whatsapp.html')


# Service Requisition Routes
@app.route('/employee/service_requisitions')
@login_required
def service_requisitions_list():
    if 'employee_id' not in session:
        flash('Access denied!', 'danger')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Get all requisitions with vehicle and requester details
        cursor.execute("""
            SELECT sr.*, 
                   v.vehicle_number, v.make, v.model,
                   e.username as requested_by_name,
                   lm.username as line_manager_name,
                   d.username as director_name
            FROM service_requisitions sr
            JOIN vehicles v ON sr.vehicle_id = v.id
            JOIN employees e ON sr.requested_by = e.id
            LEFT JOIN employees lm ON sr.line_manager_id = lm.id
            LEFT JOIN employees d ON sr.director_id = d.id
            ORDER BY sr.created_at DESC
        """)
        requisitions = cursor.fetchall()
        
        return render_template('service_requisitions_list.html', requisitions=requisitions)
    finally:
        cursor.close()
        conn.close()


@app.route('/employee/create_service_requisition', methods=['GET', 'POST'])
@login_required
def create_service_requisition():
    if 'employee_id' not in session:
        flash('Access denied!', 'danger')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        vehicle_id = request.form.get('vehicle_id')
        current_mileage = request.form.get('current_mileage')
        work_description = request.form.get('work_description')
        service_history = request.form.get('service_history')
        
        if not all([vehicle_id, work_description]):
            flash('Vehicle and work description are required!', 'danger')
            return redirect(url_for('create_service_requisition'))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Get vehicle details
            cursor.execute("SELECT vehicle_number, make, model FROM vehicles WHERE id = %s", (vehicle_id,))
            vehicle = cursor.fetchone()
            
            if not vehicle:
                flash('Vehicle not found!', 'danger')
                return redirect(url_for('create_service_requisition'))
            
            # Generate requisition number
            cursor.execute("SELECT COUNT(*) as count FROM service_requisitions")
            count = cursor.fetchone()['count']
            requisition_number = f"SR-{count + 1:05d}"
            
            # Insert requisition
            cursor.execute("""
                INSERT INTO service_requisitions 
                (requisition_number, vehicle_id, vehicle_reg_number, vehicle_make, 
                 vehicle_model, current_mileage, work_description, requested_by, service_history)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (requisition_number, vehicle_id, vehicle['vehicle_number'], 
                  vehicle['make'], vehicle['model'], 
                  current_mileage if current_mileage else None,
                  work_description, session['employee_id'], service_history))
            
            conn.commit()
            flash('Service requisition created successfully!', 'success')
            return redirect(url_for('service_requisitions_list'))
        except Exception as e:
            conn.rollback()
            flash(f'Error creating requisition: {str(e)}', 'danger')
            return redirect(url_for('create_service_requisition'))
        finally:
            cursor.close()
            conn.close()
    
    # GET request - show form
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Get all vehicles
        cursor.execute("SELECT * FROM vehicles ORDER BY vehicle_number")
        vehicles = cursor.fetchall()
        
        return render_template('create_service_requisition.html', vehicles=vehicles)
    finally:
        cursor.close()
        conn.close()


@app.route('/employee/view_service_requisition/<int:requisition_id>')
@login_required
def view_service_requisition(requisition_id):
    if 'employee_id' not in session:
        flash('Access denied!', 'danger')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Get requisition details
        cursor.execute("""
            SELECT sr.*, 
                   v.vehicle_number, v.make, v.model, v.year, v.color, v.vehicle_type,
                   e.username as requested_by_name, e.email as requester_email,
                   e.department as requester_department,
                   lm.username as line_manager_name,
                   d.username as director_name
            FROM service_requisitions sr
            JOIN vehicles v ON sr.vehicle_id = v.id
            JOIN employees e ON sr.requested_by = e.id
            LEFT JOIN employees lm ON sr.line_manager_id = lm.id
            LEFT JOIN employees d ON sr.director_id = d.id
            WHERE sr.id = %s
        """, (requisition_id,))
        requisition = cursor.fetchone()
        
        if not requisition:
            flash('Requisition not found!', 'danger')
            return redirect(url_for('service_requisitions_list'))
        
        # Get service history for the vehicle
        cursor.execute("""
            SELECT * FROM service_maintenance 
            WHERE vehicle_id = %s 
            ORDER BY service_date DESC 
            LIMIT 5
        """, (requisition['vehicle_id'],))
        service_history = cursor.fetchall()
        
        return render_template('view_service_requisition.html', 
                             requisition=requisition,
                             service_history=service_history)
    finally:
        cursor.close()
        conn.close()


@app.route('/employee/review_service_requisition/<int:requisition_id>', methods=['POST'])
@login_required
def review_service_requisition(requisition_id):
    if 'employee_id' not in session:
        flash('Access denied!', 'danger')
        return redirect(url_for('index'))
    
    action = request.form.get('action')  # 'approve' or 'reject'
    comments = request.form.get('comments')
    review_type = request.form.get('review_type')  # 'line_manager' or 'director'
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        if review_type == 'line_manager':
            status = 'approved' if action == 'approve' else 'rejected'
            cursor.execute("""
                UPDATE service_requisitions 
                SET line_manager_id = %s,
                    line_manager_status = %s,
                    line_manager_comments = %s,
                    line_manager_reviewed_at = NOW(),
                    overall_status = %s
                WHERE id = %s
            """, (session['employee_id'], status, comments, 
                  'awaiting_director' if status == 'approved' else 'rejected',
                  requisition_id))
            
            flash(f'Requisition {status} by line manager!', 'success')
            
        elif review_type == 'director':
            # Check if line manager approved
            cursor.execute("""
                SELECT line_manager_status FROM service_requisitions WHERE id = %s
            """, (requisition_id,))
            req = cursor.fetchone()
            
            if req and req['line_manager_status'] != 'approved':
                flash('Line manager must approve first!', 'warning')
                return redirect(url_for('view_service_requisition', requisition_id=requisition_id))
            
            status = 'approved' if action == 'approve' else 'rejected'
            cursor.execute("""
                UPDATE service_requisitions 
                SET director_id = %s,
                    director_status = %s,
                    director_comments = %s,
                    director_approved_at = NOW(),
                    overall_status = %s
                WHERE id = %s
            """, (session['employee_id'], status, comments, status, requisition_id))
            
            flash(f'Requisition {status} by director!', 'success')
        
        conn.commit()
        return redirect(url_for('view_service_requisition', requisition_id=requisition_id))
        
    except Exception as e:
        conn.rollback()
        flash(f'Error reviewing requisition: {str(e)}', 'danger')
        return redirect(url_for('view_service_requisition', requisition_id=requisition_id))
    finally:
        cursor.close()
        conn.close()


# Employee Management Routes (Admin)
@app.route('/employee/manage_employees')
@login_required
def manage_employees():
    if 'employee_id' not in session:
        flash('Access denied!', 'danger')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Check if current user has admin role
        cursor.execute("SELECT role FROM employees WHERE id = %s", (session['employee_id'],))
        current_user = cursor.fetchone()
        
        if not current_user or current_user['role'] != 'admin':
            flash('Admin access required!', 'danger')
            return redirect(url_for('employee_dashboard'))
        
        # Get all employees
        cursor.execute("""
            SELECT id, employee_id, username, email, department, position, 
                   role, status, phone, created_at, updated_at
            FROM employees
            ORDER BY created_at DESC
        """)
        employees = cursor.fetchall()
        
        return render_template('manage_employees.html', employees=employees)
    finally:
        cursor.close()
        conn.close()


@app.route('/employee/edit_employee/<int:emp_id>', methods=['GET', 'POST'])
@login_required
def edit_employee(emp_id):
    if 'employee_id' not in session:
        flash('Access denied!', 'danger')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Check if current user has admin role
        cursor.execute("SELECT role FROM employees WHERE id = %s", (session['employee_id'],))
        current_user = cursor.fetchone()
        
        if not current_user or current_user['role'] != 'admin':
            flash('Admin access required!', 'danger')
            return redirect(url_for('employee_dashboard'))
        
        if request.method == 'POST':
            username = request.form.get('username')
            email = request.form.get('email')
            department = request.form.get('department')
            position = request.form.get('position')
            role = request.form.get('role')
            status = request.form.get('status')
            phone = request.form.get('phone')
            
            if not all([username, email, role, status]):
                flash('Username, email, role and status are required!', 'danger')
                return redirect(url_for('edit_employee', emp_id=emp_id))
            
            # Check for duplicate username (excluding current employee)
            cursor.execute("SELECT id FROM employees WHERE username = %s AND id != %s", (username, emp_id))
            if cursor.fetchone():
                flash('Username already exists!', 'danger')
                return redirect(url_for('edit_employee', emp_id=emp_id))
            
            # Check for duplicate email (excluding current employee)
            cursor.execute("SELECT id FROM employees WHERE email = %s AND id != %s", (email, emp_id))
            if cursor.fetchone():
                flash('Email already exists!', 'danger')
                return redirect(url_for('edit_employee', emp_id=emp_id))
            
            cursor.execute("""
                UPDATE employees 
                SET username = %s, email = %s, department = %s, position = %s,
                    role = %s, status = %s, phone = %s
                WHERE id = %s
            """, (username, email, department, position, role, status, phone, emp_id))
            
            conn.commit()
            flash('Employee updated successfully!', 'success')
            return redirect(url_for('manage_employees'))
        
        # GET request - fetch employee details
        cursor.execute("SELECT * FROM employees WHERE id = %s", (emp_id,))
        employee = cursor.fetchone()
        
        if not employee:
            flash('Employee not found!', 'danger')
            return redirect(url_for('manage_employees'))
        
        return render_template('edit_employee.html', employee=employee)
        
    except Exception as e:
        conn.rollback()
        flash(f'Error updating employee: {str(e)}', 'danger')
        return redirect(url_for('edit_employee', emp_id=emp_id))
    finally:
        cursor.close()
        conn.close()


@app.route('/employee/delete_employee/<int:emp_id>', methods=['POST'])
@login_required
def delete_employee(emp_id):
    if 'employee_id' not in session:
        flash('Access denied!', 'danger')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Check if current user has admin role
        cursor.execute("SELECT role FROM employees WHERE id = %s", (session['employee_id'],))
        current_user = cursor.fetchone()
        
        if not current_user or current_user['role'] != 'admin':
            flash('Admin access required!', 'danger')
            return redirect(url_for('employee_dashboard'))
        
        # Prevent deleting yourself
        if emp_id == session['employee_id']:
            flash('You cannot delete your own account!', 'danger')
            return redirect(url_for('manage_employees'))
        
        cursor.execute("DELETE FROM employees WHERE id = %s", (emp_id,))
        conn.commit()
        
        flash('Employee deleted successfully!', 'success')
        return redirect(url_for('manage_employees'))
        
    except Exception as e:
        conn.rollback()
        flash(f'Error deleting employee: {str(e)}', 'danger')
        return redirect(url_for('manage_employees'))
    finally:
        cursor.close()
        conn.close()


@app.route('/employee/toggle_employee_status/<int:emp_id>', methods=['POST'])
@login_required
def toggle_employee_status(emp_id):
    if 'employee_id' not in session:
        flash('Access denied!', 'danger')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Check if current user has admin role
        cursor.execute("SELECT role FROM employees WHERE id = %s", (session['employee_id'],))
        current_user = cursor.fetchone()
        
        if not current_user or current_user['role'] != 'admin':
            flash('Admin access required!', 'danger')
            return redirect(url_for('employee_dashboard'))
        
        # Get current status
        cursor.execute("SELECT status FROM employees WHERE id = %s", (emp_id,))
        employee = cursor.fetchone()
        
        if not employee:
            flash('Employee not found!', 'danger')
            return redirect(url_for('manage_employees'))
        
        # Toggle status
        new_status = 'inactive' if employee['status'] == 'active' else 'active'
        cursor.execute("UPDATE employees SET status = %s WHERE id = %s", (new_status, emp_id))
        conn.commit()
        
        flash(f'Employee status changed to {new_status}!', 'success')
        return redirect(url_for('manage_employees'))
        
    except Exception as e:
        conn.rollback()
        flash(f'Error toggling employee status: {str(e)}', 'danger')
        return redirect(url_for('manage_employees'))
    finally:
        cursor.close()
        conn.close()


# Roles and Permissions Management
@app.route('/employee/manage_roles')
@login_required
def manage_roles():
    if 'employee_id' not in session:
        flash('Access denied!', 'danger')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Check if current user has admin role
        cursor.execute("SELECT role FROM employees WHERE id = %s", (session['employee_id'],))
        current_user = cursor.fetchone()
        
        if not current_user or current_user['role'] != 'admin':
            flash('Admin access required!', 'danger')
            return redirect(url_for('employee_dashboard'))
        
        # Get all roles with permission counts
        cursor.execute("""
            SELECT r.id, r.role_key, r.role_name, r.description, r.is_system_role,
                   COUNT(rp.permission_id) as permission_count,
                   COUNT(DISTINCT e.id) as employee_count
            FROM roles r
            LEFT JOIN role_permissions rp ON r.id = rp.role_id
            LEFT JOIN employees e ON e.role = r.role_key
            GROUP BY r.id
            ORDER BY r.role_name
        """)
        roles = cursor.fetchall()
        
        return render_template('manage_roles.html', roles=roles)
    finally:
        cursor.close()
        conn.close()


@app.route('/employee/edit_role/<int:role_id>', methods=['GET', 'POST'])
@login_required
def edit_role(role_id):
    if 'employee_id' not in session:
        flash('Access denied!', 'danger')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Check if current user has admin role
        cursor.execute("SELECT role FROM employees WHERE id = %s", (session['employee_id'],))
        current_user = cursor.fetchone()
        
        if not current_user or current_user['role'] != 'admin':
            flash('Admin access required!', 'danger')
            return redirect(url_for('employee_dashboard'))
        
        if request.method == 'POST':
            permission_ids = request.form.getlist('permissions')
            
            # Delete existing permissions
            cursor.execute("DELETE FROM role_permissions WHERE role_id = %s", (role_id,))
            
            # Insert new permissions
            for perm_id in permission_ids:
                cursor.execute("""
                    INSERT INTO role_permissions (role_id, permission_id)
                    VALUES (%s, %s)
                """, (role_id, perm_id))
            
            conn.commit()
            flash('Role permissions updated successfully!', 'success')
            return redirect(url_for('manage_roles'))
        
        # GET request
        cursor.execute("SELECT * FROM roles WHERE id = %s", (role_id,))
        role = cursor.fetchone()
        
        if not role:
            flash('Role not found!', 'danger')
            return redirect(url_for('manage_roles'))
        
        # Get all permissions grouped by module
        cursor.execute("""
            SELECT id, permission_key, permission_name, description, module
            FROM permissions
            ORDER BY module, permission_name
        """)
        all_permissions = cursor.fetchall()
        
        # Group by module
        permissions_by_module = {}
        for perm in all_permissions:
            module = perm['module'] or 'Other'
            if module not in permissions_by_module:
                permissions_by_module[module] = []
            permissions_by_module[module].append(perm)
        
        # Get current role permissions
        cursor.execute("""
            SELECT permission_id 
            FROM role_permissions 
            WHERE role_id = %s
        """, (role_id,))
        current_permissions = [p['permission_id'] for p in cursor.fetchall()]
        
        return render_template('edit_role.html', 
                             role=role,
                             permissions_by_module=permissions_by_module,
                             current_permissions=current_permissions)
        
    except Exception as e:
        conn.rollback()
        flash(f'Error updating role: {str(e)}', 'danger')
        return redirect(url_for('manage_roles'))
    finally:
        cursor.close()
        conn.close()


@app.route('/employee/create_role', methods=['GET', 'POST'])
@login_required
def create_role():
    if 'employee_id' not in session:
        flash('Access denied!', 'danger')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Check if current user has admin role
        cursor.execute("SELECT role FROM employees WHERE id = %s", (session['employee_id'],))
        current_user = cursor.fetchone()
        
        if not current_user or current_user['role'] != 'admin':
            flash('Admin access required!', 'danger')
            return redirect(url_for('employee_dashboard'))
        
        if request.method == 'POST':
            role_key = request.form.get('role_key')
            role_name = request.form.get('role_name')
            description = request.form.get('description')
            permission_ids = request.form.getlist('permissions')
            
            if not all([role_key, role_name]):
                flash('Role key and name are required!', 'danger')
                return redirect(url_for('create_role'))
            
            # Check for duplicate role key
            cursor.execute("SELECT id FROM roles WHERE role_key = %s", (role_key,))
            if cursor.fetchone():
                flash('Role key already exists!', 'danger')
                return redirect(url_for('create_role'))
            
            # Insert role
            cursor.execute("""
                INSERT INTO roles (role_key, role_name, description, is_system_role)
                VALUES (%s, %s, %s, FALSE)
            """, (role_key, role_name, description))
            
            role_id = cursor.lastrowid
            
            # Insert permissions
            for perm_id in permission_ids:
                cursor.execute("""
                    INSERT INTO role_permissions (role_id, permission_id)
                    VALUES (%s, %s)
                """, (role_id, perm_id))
            
            conn.commit()
            flash('Role created successfully!', 'success')
            return redirect(url_for('manage_roles'))
        
        # GET request - show form
        cursor.execute("""
            SELECT id, permission_key, permission_name, description, module
            FROM permissions
            ORDER BY module, permission_name
        """)
        all_permissions = cursor.fetchall()
        
        # Group by module
        permissions_by_module = {}
        for perm in all_permissions:
            module = perm['module'] or 'Other'
            if module not in permissions_by_module:
                permissions_by_module[module] = []
            permissions_by_module[module].append(perm)
        
        return render_template('create_role.html', 
                             permissions_by_module=permissions_by_module)
        
    except Exception as e:
        conn.rollback()
        flash(f'Error creating role: {str(e)}', 'danger')
        return redirect(url_for('create_role'))
    finally:
        cursor.close()
        conn.close()


@app.route('/employee/delete_role/<int:role_id>', methods=['POST'])
@login_required
def delete_role(role_id):
    if 'employee_id' not in session:
        flash('Access denied!', 'danger')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Check if current user has admin role
        cursor.execute("SELECT role FROM employees WHERE id = %s", (session['employee_id'],))
        current_user = cursor.fetchone()
        
        if not current_user or current_user['role'] != 'admin':
            flash('Admin access required!', 'danger')
            return redirect(url_for('employee_dashboard'))
        
        # Check if it's a system role
        cursor.execute("SELECT is_system_role, role_key FROM roles WHERE id = %s", (role_id,))
        role = cursor.fetchone()
        
        if not role:
            flash('Role not found!', 'danger')
            return redirect(url_for('manage_roles'))
        
        if role['is_system_role']:
            flash('Cannot delete system roles!', 'danger')
            return redirect(url_for('manage_roles'))
        
        # Check if any employees have this role
        cursor.execute("SELECT COUNT(*) as count FROM employees WHERE role = %s", (role['role_key'],))
        count = cursor.fetchone()['count']
        
        if count > 0:
            flash(f'Cannot delete role: {count} employee(s) currently have this role!', 'danger')
            return redirect(url_for('manage_roles'))
        
        # Delete role (permissions will be deleted by CASCADE)
        cursor.execute("DELETE FROM roles WHERE id = %s", (role_id,))
        conn.commit()
        
        flash('Role deleted successfully!', 'success')
        return redirect(url_for('manage_roles'))
        
    except Exception as e:
        conn.rollback()
        flash(f'Error deleting role: {str(e)}', 'danger')
        return redirect(url_for('manage_roles'))
    finally:
        cursor.close()
        conn.close()

# Data Import Routes
@app.route('/data-import')
def data_import():
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    return render_template('data_import.html')

@app.route('/data-import/download-template/<data_type>')
def download_import_template(data_type):
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask import send_file
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    
    # Style for header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    if data_type == 'vehicles':
        from openpyxl.worksheet.datavalidation import DataValidation
        
        ws.title = 'Vehicles Template'
        headers = ['Vehicle Number*', 'Make*', 'Model*', 'Year*', 'Color', 
                   'Vehicle Type', 'Status', 'Mileage', 'Last Service Date (YYYY-MM-DD)', 
                   'Expected Fuel Consumption (km/L)', 'Notes']
        ws.append(headers)
        
        # Add example row
        ws.append(['V001', 'Toyota', 'Corolla', '2020', 'White', 
                   'Sedan', 'available', '50000', '2024-01-15', '14.5', 'Fleet vehicle'])
        
        # Add data validation for Status
        status_list = 'available,in_use,maintenance,out_of_service'
        status_dv = DataValidation(type="list", formula1=f'"{status_list}"', allow_blank=True)
        status_dv.prompt = 'Select status from the dropdown'
        status_dv.promptTitle = 'Vehicle Status'
        ws.add_data_validation(status_dv)
        status_dv.add(f'G2:G1001')
        
        # Add data validation for Fuel Type
        fuel_type_list = 'Petrol,Diesel,Electric,Hybrid'
        fuel_type_dv = DataValidation(type="list", formula1=f'"{fuel_type_list}"', allow_blank=True)
        fuel_type_dv.prompt = 'Select fuel type from the dropdown'
        fuel_type_dv.promptTitle = 'Fuel Type'
        ws.add_data_validation(fuel_type_dv)
        fuel_type_dv.add(f'H2:H1001')
        
        # Add data validation for Transmission
        transmission_list = 'Automatic,Manual'
        transmission_dv = DataValidation(type="list", formula1=f'"{transmission_list}"', allow_blank=True)
        transmission_dv.prompt = 'Select transmission from the dropdown'
        transmission_dv.promptTitle = 'Transmission'
        ws.add_data_validation(transmission_dv)
        transmission_dv.add(f'I2:I1001')
        
        # Add notes
        ws.append([])
        ws.append(['Notes:'])
        ws.append(['* Required fields'])
        ws.append(['Status dropdown: Available, In Use, Maintenance, Out of Service'])
        ws.append(['Fuel Type dropdown: Petrol, Diesel, Electric, Hybrid'])
        ws.append(['Transmission dropdown: Automatic, Manual'])
        
    elif data_type == 'employees':
        from openpyxl.worksheet.datavalidation import DataValidation
        
        ws.title = 'Employees Template'
        headers = ['Username*', 'Email*', 'Password*', 'Employee ID*', 'Phone', 
                   'Department', 'Position', 'Status']
        ws.append(headers)
        
        # Add example row
        ws.append(['john.doe', 'john@example.com', 'password123', 'EMP001', 
                   '+1234567890', 'Operations', 'Driver', 'active'])
        
        # Add data validation for Status
        status_list = 'active,inactive'
        status_dv = DataValidation(type="list", formula1=f'"{status_list}"', allow_blank=True)
        status_dv.prompt = 'Select status from the dropdown'
        status_dv.promptTitle = 'Employee Status'
        ws.add_data_validation(status_dv)
        status_dv.add(f'H2:H1001')
        
        # Add notes
        ws.append([])
        ws.append(['Notes:'])
        ws.append(['* Required fields'])
        ws.append(['Status dropdown: active, inactive'])
        ws.append(['Password will be hashed during import'])
        
    elif data_type == 'fuel_records':
        from openpyxl.worksheet.datavalidation import DataValidation
        
        ws.title = 'Fuel Records Template'
        headers = ['Vehicle Number*', 'Employee Username*', 'Fuel Date (YYYY-MM-DD HH:MM)*', 
                   'Fuel Amount (Liters)*', 'Fuel Cost*', 'Odometer Reading', 
                   'Fuel Type', 'Station Name', 'Payment Method']
        ws.append(headers)
        
        # Get vehicles and employees from database for dropdowns
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT vehicle_number, make, model FROM vehicles ORDER BY vehicle_number")
            vehicles = cursor.fetchall()
            
            cursor.execute("SELECT username, employee_id FROM employees ORDER BY username")
            employees = cursor.fetchall()
        finally:
            cursor.close()
            conn.close()
        
        # Create reference sheets for dropdowns
        if vehicles:
            # Create Vehicles reference sheet
            vehicles_sheet = wb.create_sheet('Vehicles_Reference')
            vehicles_sheet.append(['Vehicle Number', 'Make', 'Model'])
            vehicle_numbers = []
            for vehicle in vehicles:
                vehicles_sheet.append([vehicle['vehicle_number'], vehicle['make'], vehicle['model']])
                vehicle_numbers.append(vehicle['vehicle_number'])
            
            # Hide the reference sheet
            vehicles_sheet.sheet_state = 'hidden'
            
            # Create data validation for Vehicle Number column
            if vehicle_numbers:
                vehicle_list = ','.join(vehicle_numbers)
                vehicle_dv = DataValidation(type="list", formula1=f'"{vehicle_list}"', allow_blank=False)
                vehicle_dv.error = 'Please select a vehicle from the list'
                vehicle_dv.errorTitle = 'Invalid Vehicle'
                vehicle_dv.prompt = 'Select a vehicle from the dropdown'
                vehicle_dv.promptTitle = 'Vehicle Selection'
                ws.add_data_validation(vehicle_dv)
                # Apply to first 1000 rows
                vehicle_dv.add(f'A2:A1001')
        
        if employees:
            # Create Employees reference sheet
            employees_sheet = wb.create_sheet('Employees_Reference')
            employees_sheet.append(['Username', 'Employee ID'])
            employee_usernames = []
            for employee in employees:
                employees_sheet.append([employee['username'], employee['employee_id'] or ''])
                employee_usernames.append(employee['username'])
            
            # Hide the reference sheet
            employees_sheet.sheet_state = 'hidden'
            
            # Create data validation for Employee Username column
            if employee_usernames:
                employee_list = ','.join(employee_usernames)
                employee_dv = DataValidation(type="list", formula1=f'"{employee_list}"', allow_blank=False)
                employee_dv.error = 'Please select an employee from the list'
                employee_dv.errorTitle = 'Invalid Employee'
                employee_dv.prompt = 'Select an employee from the dropdown'
                employee_dv.promptTitle = 'Employee Selection'
                ws.add_data_validation(employee_dv)
                # Apply to first 1000 rows
                employee_dv.add(f'B2:B1001')
        
        # Create data validation for Fuel Type
        fuel_types = ['Petrol', 'Diesel', 'Electric', 'Hybrid']
        fuel_type_list = ','.join(fuel_types)
        fuel_type_dv = DataValidation(type="list", formula1=f'"{fuel_type_list}"', allow_blank=True)
        fuel_type_dv.prompt = 'Select fuel type from the dropdown'
        fuel_type_dv.promptTitle = 'Fuel Type'
        ws.add_data_validation(fuel_type_dv)
        fuel_type_dv.add(f'G2:G1001')
        
        # Add example row (only if we have vehicles and employees)
        if vehicles and employees:
            ws.append([vehicles[0]['vehicle_number'], employees[0]['username'], 
                      '2024-01-15 10:30', '45.5', '2500.00', 
                      '52000', 'Petrol', 'Shell Station', 'Company Card'])
        else:
            ws.append(['', '', '2024-01-15 10:30', '45.5', '2500.00', 
                      '52000', 'Petrol', 'Shell Station', 'Company Card'])
        
        # Add notes
        ws.append([])
        ws.append(['Notes:'])
        ws.append(['* Required fields'])
        ws.append(['Vehicle Number and Employee Username have dropdown lists - click to select'])
        ws.append(['Only vehicles and employees currently in the system are available'])
        ws.append(['Fuel Type has dropdown: Petrol, Diesel, Electric, Hybrid'])
        ws.append(['Date format: YYYY-MM-DD HH:MM (e.g., 2024-01-15 10:30)'])
        ws.append(['Download fresh template if vehicles/employees have been added recently'])
    
    # Style the header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'{data_type}_import_template.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/data-import/upload', methods=['POST'])
def upload_import_data():
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    data_type = request.form.get('data_type')
    
    if 'file' not in request.files:
        flash('No file uploaded!', 'danger')
        return redirect(url_for('data_import'))
    
    file = request.files['file']
    
    if file.filename == '':
        flash('No file selected!', 'danger')
        return redirect(url_for('data_import'))
    
    if not file.filename.endswith('.xlsx'):
        flash('Only .xlsx files are allowed!', 'danger')
        return redirect(url_for('data_import'))
    
    try:
        from openpyxl import load_workbook
        from datetime import datetime
        
        wb = load_workbook(file)
        ws = wb.active
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        success_count = 0
        error_count = 0
        errors = []
        
        # Skip header row
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        if data_type == 'vehicles':
            for idx, row in enumerate(rows, start=2):
                # Skip empty rows or notes section
                if not row[0] or str(row[0]).startswith('Notes'):
                    continue
                
                try:
                    vehicle_number, make, model, year, color, vehicle_type, status, mileage, last_service_date, expected_fuel_consumption, notes = row[:11]
                    
                    # Required fields
                    if not all([vehicle_number, make, model, year]):
                        errors.append(f"Row {idx}: Missing required fields")
                        error_count += 1
                        continue
                    
                    # Check if vehicle already exists
                    cursor.execute("SELECT id FROM vehicles WHERE vehicle_number = %s", (vehicle_number,))
                    if cursor.fetchone():
                        errors.append(f"Row {idx}: Vehicle {vehicle_number} already exists")
                        error_count += 1
                        continue
                    
                    # Parse date
                    if last_service_date:
                        if isinstance(last_service_date, str):
                            last_service_date = datetime.strptime(last_service_date, '%Y-%m-%d').date()
                    
                    cursor.execute("""
                        INSERT INTO vehicles 
                        (vehicle_number, make, model, year, color, vehicle_type, 
                         status, mileage, last_service_date, expected_fuel_consumption, notes)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (vehicle_number, make, model, year, color, vehicle_type,
                          status or 'available', mileage, last_service_date, expected_fuel_consumption, notes))
                    
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {idx}: {str(e)}")
                    error_count += 1
        
        elif data_type == 'employees':
            for idx, row in enumerate(rows, start=2):
                if not row[0] or str(row[0]).startswith('Notes'):
                    continue
                
                try:
                    username, email, password, employee_id, phone, department, position, status = row[:8]
                    
                    if not all([username, email, password, employee_id]):
                        errors.append(f"Row {idx}: Missing required fields")
                        error_count += 1
                        continue
                    
                    # Check if employee exists
                    cursor.execute("SELECT id FROM employees WHERE username = %s OR email = %s OR employee_id = %s", (username, email, employee_id))
                    if cursor.fetchone():
                        errors.append(f"Row {idx}: Employee {username}, email, or employee ID already exists")
                        error_count += 1
                        continue
                    
                    # Hash password
                    hashed_password = generate_password_hash(password)
                    
                    cursor.execute("""
                        INSERT INTO employees 
                        (username, email, password_hash, employee_id, phone, department, 
                         position, status)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """, (username, email, hashed_password, employee_id, phone, department,
                          position, status or 'active'))
                    
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {idx}: {str(e)}")
                    error_count += 1
        
        elif data_type == 'fuel_records':
            # Detect whether the `payment_method` column exists in the fuel_records table
            try:
                cursor.execute("SHOW COLUMNS FROM fuel_records LIKE 'payment_method'")
                has_payment_column = cursor.fetchone() is not None
            except Exception:
                has_payment_column = False

            for idx, row in enumerate(rows, start=2):
                if not row[0] or str(row[0]).startswith('Notes'):
                    continue
                
                try:
                    # Safely unpack up to 9 columns, padding missing values with None
                    vals = list(row) + [None] * 9
                    vehicle_number, employee_username, fuel_date, fuel_amount, fuel_cost, odometer, fuel_type, station, payment = vals[:9]

                    # Trim strings
                    if isinstance(vehicle_number, str):
                        vehicle_number = vehicle_number.strip()
                    if isinstance(employee_username, str):
                        employee_username = employee_username.strip()
                    if isinstance(fuel_type, str):
                        fuel_type = fuel_type.strip()
                    if isinstance(station, str):
                        station = station.strip()
                    if isinstance(payment, str):
                        payment = payment.strip()

                    # Sanitize numeric fields: remove currency symbols/commas and convert
                    def _to_float(val):
                        if val is None:
                            return None
                        if isinstance(val, (int, float)):
                            return float(val)
                        s = str(val).strip()
                        # Remove common currency symbols and thousands separators
                        s = s.replace('$', '').replace('€', '').replace('£', '')
                        s = s.replace(',', '')
                        if s == '':
                            return None
                        try:
                            return float(s)
                        except Exception:
                            raise ValueError(f"Invalid numeric value: {val}")

                    def _to_int(val):
                        if val is None or val == '':
                            return None
                        if isinstance(val, int):
                            return val
                        try:
                            return int(float(str(val).strip().replace(',', '')))
                        except Exception:
                            raise ValueError(f"Invalid integer value: {val}")

                    try:
                        fuel_amount = _to_float(fuel_amount)
                    except Exception as e:
                        errors.append(f"Row {idx}: fuel_amount conversion error: {e} - data: {row}")
                        error_count += 1
                        continue

                    try:
                        fuel_cost = _to_float(fuel_cost)
                    except Exception as e:
                        errors.append(f"Row {idx}: fuel_cost conversion error: {e} - data: {row}")
                        error_count += 1
                        continue

                    try:
                        odometer = _to_int(odometer)
                    except Exception as e:
                        errors.append(f"Row {idx}: odometer conversion error: {e} - data: {row}")
                        error_count += 1
                        continue

                    if not all([vehicle_number, employee_username, fuel_date, fuel_amount is not None, fuel_cost is not None]):
                        errors.append(f"Row {idx}: Missing required fields")
                        error_count += 1
                        continue
                    
                    # Get vehicle ID
                    cursor.execute("SELECT id FROM vehicles WHERE vehicle_number = %s", (vehicle_number,))
                    vehicle = cursor.fetchone()
                    if not vehicle:
                        errors.append(f"Row {idx}: Vehicle {vehicle_number} not found")
                        error_count += 1
                        continue
                    
                    # Get employee ID
                    cursor.execute("SELECT id FROM employees WHERE username = %s", (employee_username,))
                    employee = cursor.fetchone()
                    if not employee:
                        errors.append(f"Row {idx}: Employee {employee_username} not found")
                        error_count += 1
                        continue
                    
                    # Parse datetime
                    if isinstance(fuel_date, str):
                        fuel_date = datetime.strptime(fuel_date, '%Y-%m-%d %H:%M')
                    
                    if has_payment_column:
                        cursor.execute("""
                            INSERT INTO fuel_records 
                            (vehicle_id, employee_id, fuel_date, fuel_amount, fuel_cost, 
                             odometer_reading, fuel_type, station_name, payment_method)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (vehicle['id'], employee['id'], fuel_date, fuel_amount, fuel_cost,
                              odometer, fuel_type, station, payment))
                    else:
                        cursor.execute("""
                            INSERT INTO fuel_records 
                            (vehicle_id, employee_id, fuel_date, fuel_amount, fuel_cost, 
                             odometer_reading, fuel_type, station_name)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        """, (vehicle['id'], employee['id'], fuel_date, fuel_amount, fuel_cost,
                              odometer, fuel_type, station))

                    success_count += 1

                except Exception as e:
                    # Include row data to aid debugging of malformed rows or schema mismatches
                    errors.append(f"Row {idx}: {str(e)} - data: {row}")
                    error_count += 1
        
        conn.commit()
        
        # Show results
        if success_count > 0:
            flash(f'Successfully imported {success_count} records!', 'success')
        
        if error_count > 0:
            flash(f'{error_count} records failed to import.', 'warning')
            for error in errors[:10]:  # Show first 10 errors
                flash(error, 'danger')
            if len(errors) > 10:
                flash(f'... and {len(errors) - 10} more errors', 'danger')
        
        return redirect(url_for('data_import'))
        
    except Exception as e:
        flash(f'Error processing file: {str(e)}', 'danger')
        return redirect(url_for('data_import'))
    finally:
        cursor.close()
        conn.close()

# Fuel Exceptions Report
@app.route('/fuel-exceptions-report')
def fuel_exceptions_report():
    if 'employee_id' not in session:
        flash('Please log in to access this page.', 'warning')
        return redirect(url_for('employee_login'))
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Get default threshold setting
        default_threshold = float(get_setting('default_fuel_consumption_threshold', 18))
        
        # Get filters
        vehicle_filter = request.args.get('vehicle_filter', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        
        # Build query to get fuel records with consumption calculation
        query = """
            SELECT 
                fr.id,
                fr.fuel_date,
                fr.fuel_amount,
                fr.fuel_cost,
                fr.odometer_reading,
                fr.station_name,
                v.id as vehicle_id,
                v.vehicle_number,
                v.make,
                v.model,
                v.expected_fuel_consumption,
                e.username as employee_name,
                e.employee_id
            FROM fuel_records fr
            JOIN vehicles v ON fr.vehicle_id = v.id
            JOIN employees e ON fr.employee_id = e.id
            WHERE 1=1
        """
        params = []
        
        if vehicle_filter:
            query += " AND v.id = %s"
            params.append(vehicle_filter)
        
        if date_from:
            query += " AND fr.fuel_date >= %s"
            params.append(date_from)
        
        if date_to:
            query += " AND fr.fuel_date <= %s"
            params.append(date_to + ' 23:59:59')
        
        query += " ORDER BY v.vehicle_number, fr.fuel_date"
        
        cursor.execute(query, params)
        fuel_records = cursor.fetchall()
        
        # Calculate consumption for each record and identify exceptions
        exceptions = []
        vehicles_data = {}
        
        for i, record in enumerate(fuel_records):
            vehicle_id = record['vehicle_id']
            
            if vehicle_id not in vehicles_data:
                vehicles_data[vehicle_id] = []
            
            vehicles_data[vehicle_id].append(record)
        
        # Process each vehicle's records
        for vehicle_id, records in vehicles_data.items():
            # Sort by date
            records.sort(key=lambda x: x['fuel_date'])
            
            # Get expected consumption (vehicle-specific or default)
            expected_consumption = records[0]['expected_fuel_consumption'] or default_threshold
            
            for i in range(1, len(records)):
                current = records[i]
                previous = records[i-1]
                
                # Only calculate if both have odometer readings
                if current['odometer_reading'] and previous['odometer_reading']:
                    distance = float(current['odometer_reading']) - float(previous['odometer_reading'])
                    
                    if distance > 0:
                        fuel_used = float(previous['fuel_amount'])
                        actual_consumption = distance / fuel_used  # km per liter
                        
                        # Check if consumption is worse than expected (lower km/L = worse)
                        if actual_consumption < expected_consumption:
                            exception_data = {
                                'fuel_record_id': previous['id'],
                                'fuel_date': previous['fuel_date'],
                                'vehicle_number': previous['vehicle_number'],
                                'vehicle_make_model': f"{previous['make']} {previous['model']}",
                                'employee': f"{previous['employee_name']} ({previous['employee_id']})",
                                'fuel_amount': previous['fuel_amount'],
                                'fuel_cost': previous['fuel_cost'],
                                'distance': distance,
                                'actual_consumption': actual_consumption,
                                'expected_consumption': expected_consumption,
                                'difference': expected_consumption - actual_consumption,
                                'percentage_worse': ((expected_consumption - actual_consumption) / expected_consumption) * 100,
                                'station_name': previous['station_name'],
                                'odometer_from': previous['odometer_reading'],
                                'odometer_to': current['odometer_reading']
                            }
                            exceptions.append(exception_data)
        
        # Sort by percentage worse (worst first)
        exceptions.sort(key=lambda x: x['percentage_worse'], reverse=True)
        
        # Summary statistics
        summary = {
            'total_exceptions': len(exceptions),
            'total_vehicles': len(set(e['vehicle_number'] for e in exceptions)),
            'avg_percentage_worse': sum(e['percentage_worse'] for e in exceptions) / len(exceptions) if exceptions else 0,
            'worst_consumption': min(exceptions, key=lambda x: x['actual_consumption']) if exceptions else None
        }
        
        # Get vehicles list for filter
        cursor.execute("SELECT id, vehicle_number, make, model FROM vehicles ORDER BY vehicle_number")
        vehicles = cursor.fetchall()
        
        return render_template('fuel_exceptions_report.html',
                             exceptions=exceptions,
                             summary=summary,
                             vehicles=vehicles,
                             vehicle_filter=vehicle_filter,
                             date_from=date_from,
                             date_to=date_to,
                             default_threshold=default_threshold)
    
    finally:
        cursor.close()
        conn.close()


# API endpoints for analytics report builder
@app.route('/api/vehicles')
@login_required
def api_vehicles():
    """Get list of all vehicles for report builder dropdowns"""
    if 'employee_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT id, vehicle_number, make, model, year, status
            FROM vehicles
            ORDER BY vehicle_number
        """)
        vehicles = cursor.fetchall()
        
        return jsonify([{
            'id': v['id'],
            'vehicle_number': v['vehicle_number'],
            'make': v['make'],
            'model': v['model'],
            'year': v['year'],
            'status': v['status'],
            'display_name': f"{v['vehicle_number']} - {v['make']} {v['model']}"
        } for v in vehicles])
    
    finally:
        cursor.close()
        conn.close()


@app.route('/api/employees')
@login_required
def api_employees():
    """Get list of all employees for report builder dropdowns"""
    if 'employee_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT id, username, email, employee_id, department, position, status
            FROM employees
            ORDER BY username
        """)
        employees = cursor.fetchall()
        
        return jsonify([{
            'id': e['id'],
            'username': e['username'],
            'email': e['email'],
            'employee_id': e['employee_id'],
            'department': e['department'],
            'position': e['position'],
            'status': e['status'],
            'display_name': f"{e['username']} ({e['employee_id'] if e['employee_id'] else 'No ID'})"
        } for e in employees])
    
    finally:
        cursor.close()
        conn.close()
